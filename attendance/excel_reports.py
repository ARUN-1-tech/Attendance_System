import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.db.models import Q
from django.utils import timezone
import datetime
from accounts.models import Student, Class, Subject
from attendance.models import Attendance, filter_active_attendance
from timetable.models import Schedule

def generate_attendance_excel_report(
    report_mode='day',
    class_id=None,
    subject_id=None,
    from_date_str=None,
    to_date_str=None,
    tutor_user=None,
    requested_by_user=None,
    report_type='class',
    student_id=None,
    year=None
):
    """
    Generates a beautifully styled openpyxl Workbook for Day-wise or Subject-wise attendance.
    Returns the openpyxl Workbook object.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Report"
    ws.views.sheetView[0].showGridLines = True

    # Parse dates
    from_date = None
    to_date = None
    if from_date_str:
        try:
            from_date = datetime.datetime.strptime(from_date_str, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            pass
    if to_date_str:
        try:
            to_date = datetime.datetime.strptime(to_date_str, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            pass

    if report_mode == 'day' and not from_date:
        from_date = timezone.localdate()
        to_date = from_date

    # Dynamic defaults for dates if still not provided
    if not from_date or not to_date:
        q_filter = {}
        if class_id:
            q_filter['student__student_class_id'] = class_id
        if subject_id:
            q_filter['schedule__subject_id'] = subject_id
        if tutor_user:
            q_filter['student__tutor'] = tutor_user

        if not from_date:
            first_record = Attendance.objects.filter(**q_filter).order_by('date').first()
            if first_record:
                from_date = first_record.date
            else:
                from_date = timezone.localdate()

        if not to_date:
            last_record = Attendance.objects.filter(**q_filter).order_by('-date').first()
            if last_record:
                to_date = last_record.date
            else:
                to_date = from_date or timezone.localdate()

    if from_date and to_date and from_date > to_date:
        from_date, to_date = to_date, from_date

    selected_subject = None
    if subject_id:
        try:
            selected_subject = Subject.objects.get(pk=subject_id)
        except Subject.DoesNotExist:
            pass

    selected_class = None
    if class_id:
        try:
            selected_class = Class.objects.get(pk=class_id)
        except Class.DoesNotExist:
            pass

    # Determine Student Roster
    if selected_subject:
        students_qs = selected_subject.get_enrolled_students().select_related('user', 'student_class')
        if selected_class:
            students_qs = students_qs.filter(student_class=selected_class)
    elif selected_class:
        students_qs = Student.objects.filter(student_class=selected_class).select_related('user', 'student_class')
    else:
        students_qs = Student.objects.all().select_related('user', 'student_class')

    if tutor_user:
        students_qs = students_qs.filter(tutor=tutor_user)

    if report_type == 'student' and student_id:
        from django.db.models import Q
        students_qs = students_qs.filter(Q(user__username=student_id) | Q(reg_no=student_id))
        try:
            single_student = Student.objects.select_related('student_class').get(Q(user__username=student_id) | Q(reg_no=student_id))
            if not selected_class:
                selected_class = single_student.student_class
        except Student.DoesNotExist:
            pass
    elif report_type == 'tutored':
        students_qs = students_qs.filter(tutor=requested_by_user)

    if year:
        try:
            students_qs = students_qs.filter(student_class__year=int(year))
        except (ValueError, TypeError):
            pass

    students = list(students_qs.order_by('student_class__name', 'student_class__section', 'reg_no', 'user__username'))

    # Fetch all approved and verified OD dates for the students in this report
    from leave.models import Leave
    verified_od_dates = Leave.objects.filter(
        student__in=students,
        leave_type='OD',
        final_status='Approved',
        certificate_verified=True
    ).values_list('student_id', 'date')
    verified_od_map = set((student_id, dt) for student_id, dt in verified_od_dates)

    # Styling Palette
    header_fill = PatternFill(start_color='0F172A', end_color='0F172A', fill_type='solid') # Navy
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')

    title_font = Font(name='Calibri', size=14, bold=True, color='0F172A')
    subtitle_font = Font(name='Calibri', size=11, bold=True, color='475569')
    meta_font = Font(name='Calibri', size=10, italic=True, color='64748B')
    bold_font = Font(name='Calibri', size=11, bold=True, color='0F172A')
    regular_font = Font(name='Calibri', size=11, color='1E293B')

    # Status Fills & Fonts
    present_fill = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid')
    present_font = Font(name='Calibri', size=11, bold=True, color='15803D')

    absent_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
    absent_font = Font(name='Calibri', size=11, bold=True, color='B91C1C')

    od_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
    od_font = Font(name='Calibri', size=11, bold=True, color='B45309')

    zebra_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
    summary_row_fill = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')

    thin_side = Side(style='thin', color='CBD5E1')
    grid_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')

    # Write Institutional Header Block
    ws['A1'] = "DR. NGP INSTITUTE OF TECHNOLOGY (AUTONOMOUS)"
    ws['A1'].font = title_font
    ws['A2'] = f"DEPARTMENT OF {(selected_class.department.name.upper() if selected_class and selected_class.department else 'COMPUTER SCIENCE AND ENGINEERING')}"
    ws['A2'].font = subtitle_font

    report_title = "DAY-WISE ATTENDANCE REPORT" if report_mode == 'day' else "SUBJECT-WISE ATTENDANCE REPORT"
    ws['A3'] = report_title
    ws['A3'].font = Font(name='Calibri', size=12, bold=True, color='1D4ED8')

    date_str_formatted = f"{from_date.strftime('%d-%b-%Y')} to {to_date.strftime('%d-%b-%Y')}" if from_date != to_date else from_date.strftime('%d-%b-%Y')
    class_str = selected_class.display_name if selected_class else "All Classes"
    subject_str = f" | Subject: {selected_subject.code} - {selected_subject.name}" if selected_subject else ""
    gen_by_str = f" | Generated By: {requested_by_user.get_full_name() or requested_by_user.username}" if requested_by_user else ""
    
    ws['A4'] = f"Class: {class_str}{subject_str} | Date Range: {date_str_formatted}{gen_by_str} | Generated: {timezone.now().strftime('%d-%b-%Y %I:%M %p')}"
    ws['A4'].font = meta_font

    current_row = 6

    # ==========================================
    # DAY-WISE REPORT GENERATION
    # ==========================================
    if report_mode == 'day':
        # Collect distinct dates in range where attendance actually exists for these students
        db_dates = Attendance.objects.filter(
            student__in=students,
            date__gte=from_date,
            date__lte=to_date
        ).values_list('date', flat=True).distinct()
        target_dates = sorted(list(set(db_dates)))

        # Base Columns
        headers = ["S.No", "Student Name", "Register Number", "Roll Number", "Class"]
        
        # Add Dynamic Date Columns
        for d in target_dates:
            headers.append(d.strftime('%d-%b-%Y'))

        # Add End Summary Columns
        headers.extend(["Total Days", "Present", "Absent", "OD", "Attendance %"])

        # Write Header Row
        for col_idx, text in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=text)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align_center if col_idx != 2 else align_left
            cell.border = grid_border

        ws.row_dimensions[current_row].height = 28
        current_row += 1

        # Query all attendance records for students in date range
        atts = Attendance.objects.filter(
            student__in=students,
            date__gte=from_date,
            date__lte=to_date
        ).select_related('student', 'schedule')
        atts = filter_active_attendance(atts)

        # Map (student_id, date) -> list of period statuses
        att_map = {}
        for a in atts:
            key = (a.student_id, a.date)
            if key not in att_map:
                att_map[key] = []
            att_map[key].append(a.status)

        class_total_p = 0
        class_total_a = 0
        class_total_od = 0
        class_total_possible_days = 0

        # Initialize daily summaries
        day_p_counts = {d: 0.0 for d in target_dates}
        day_a_counts = {d: 0.0 for d in target_dates}
        day_od_counts = {d: 0.0 for d in target_dates}

        # Populate Student Rows
        for idx, student in enumerate(students, start=1):
            row_num = current_row
            ws.cell(row=row_num, column=1, value=idx).alignment = align_center
            ws.cell(row=row_num, column=2, value=student.user.get_full_name() or student.user.username).alignment = align_left
            ws.cell(row=row_num, column=3, value=student.reg_no or student.user.username).alignment = align_center
            ws.cell(row=row_num, column=4, value=student.roll_no or '-').alignment = align_center
            ws.cell(row=row_num, column=5, value=student.student_class.display_name if student.student_class else '-').alignment = align_center

            for c in range(1, 6):
                cell = ws.cell(row=row_num, column=c)
                cell.font = regular_font
                cell.border = grid_border
                if idx % 2 == 0:
                    cell.fill = zebra_fill

            # Day Status Evaluation per Date
            st_p = 0
            st_a = 0
            st_od = 0
            st_verified_od = 0
            st_total_days = len(target_dates)

            for d_idx, d_obj in enumerate(target_dates, start=6):
                statuses = att_map.get((student.user_id, d_obj), [])
                is_verified_od_day = (student.user_id, d_obj) in verified_od_map
                
                # Determine daily overall status
                if not statuses:
                    day_status = "A" # Default absent if unrecorded
                    st_a += 1
                else:
                    p_count = statuses.count('Present')
                    a_count = statuses.count('Absent')
                    od_count = statuses.count('OD')
                    leave_count = statuses.count('Leave')

                    if a_count == 0 and od_count == 0 and leave_count == 0 and p_count > 0:
                        day_status = "P"
                        st_p += 1
                    elif od_count > 0 and a_count == 0 and leave_count == 0:
                        day_status = "OD"
                        st_od += 1
                        if is_verified_od_day:
                            st_verified_od += 1
                    elif a_count > 0 or leave_count > 0:
                        if p_count > 0:
                            day_status = "HD"
                            st_p += 0.5
                            st_a += 0.5
                        else:
                            day_status = "A"
                            st_a += 1
                    else:
                        day_status = "A"
                        st_a += 1

                # Update daily count summaries
                if day_status == "P":
                    day_p_counts[d_obj] += 1
                elif day_status == "A":
                    day_a_counts[d_obj] += 1
                elif day_status == "OD":
                    day_od_counts[d_obj] += 1
                elif day_status == "HD":
                    day_p_counts[d_obj] += 0.5
                    day_a_counts[d_obj] += 0.5

                cell = ws.cell(row=row_num, column=d_idx, value=day_status)
                cell.alignment = align_center
                cell.border = grid_border

                if day_status == "P":
                    cell.fill = present_fill
                    cell.font = present_font
                elif day_status == "A":
                    cell.fill = absent_fill
                    cell.font = absent_font
                elif day_status in ["OD", "HD"]:
                    cell.fill = od_fill
                    cell.font = od_font
                else:
                    cell.font = regular_font

            # End Summary Calculations
            st_effective_present = st_p + st_verified_od
            att_pct = round((st_effective_present / st_total_days * 100), 1) if st_total_days > 0 else 0.0

            sum_col_start = 6 + len(target_dates)
            
            c_tot = ws.cell(row=row_num, column=sum_col_start, value=st_total_days)
            c_p = ws.cell(row=row_num, column=sum_col_start+1, value=st_p)
            c_a = ws.cell(row=row_num, column=sum_col_start+2, value=st_a)
            c_od = ws.cell(row=row_num, column=sum_col_start+3, value=st_od)
            c_pct = ws.cell(row=row_num, column=sum_col_start+4, value=f"{att_pct}%")

            for sc_idx, cell_obj in enumerate([c_tot, c_p, c_a, c_od, c_pct], start=sum_col_start):
                cell_obj.alignment = align_center
                cell_obj.font = bold_font
                cell_obj.border = grid_border
                if idx % 2 == 0:
                    cell_obj.fill = zebra_fill

            class_total_p += st_effective_present
            class_total_a += st_a
            class_total_od += st_od
            class_total_possible_days += st_total_days

            ws.row_dimensions[row_num].height = 20
            current_row += 1

        # Add Class Summary Row at Bottom
        if len(students) > 0:
            ws.cell(row=current_row, column=1, value="").border = grid_border
            ws.cell(row=current_row, column=2, value="CLASS AVERAGE SUMMARY").font = bold_font
            ws.cell(row=current_row, column=2).alignment = align_left
            ws.cell(row=current_row, column=2).border = grid_border

            for c in range(3, 6):
                cell = ws.cell(row=current_row, column=c, value="-")
                cell.alignment = align_center
                cell.font = bold_font
                cell.border = grid_border
                cell.fill = summary_row_fill

            for d_idx, d_obj in enumerate(target_dates, start=6):
                p_sum = day_p_counts[d_obj]
                a_sum = day_a_counts[d_obj]
                od_sum = day_od_counts[d_obj]

                p_str = int(p_sum) if p_sum.is_integer() else p_sum
                a_str = int(a_sum) if a_sum.is_integer() else a_sum
                od_str = int(od_sum) if od_sum.is_integer() else od_sum

                sum_text = f"P:{p_str}\nA:{a_str}\nO:{od_str}"

                cell = ws.cell(row=current_row, column=d_idx, value=sum_text)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.font = Font(name='Calibri', size=9, bold=True, color='475569')
                cell.border = grid_border
                cell.fill = summary_row_fill

            sum_col_start = 6 + len(target_dates)
            avg_pct = round((class_total_p / class_total_possible_days * 100), 1) if class_total_possible_days > 0 else 0.0

            c_tot = ws.cell(row=current_row, column=sum_col_start, value=class_total_possible_days)
            c_p = ws.cell(row=current_row, column=sum_col_start+1, value=class_total_p)
            c_a = ws.cell(row=current_row, column=sum_col_start+2, value=class_total_a)
            c_od = ws.cell(row=current_row, column=sum_col_start+3, value=class_total_od)
            c_pct = ws.cell(row=current_row, column=sum_col_start+4, value=f"{avg_pct}%")

            for cell_obj in [c_tot, c_p, c_a, c_od, c_pct]:
                cell_obj.alignment = align_center
                cell_obj.font = Font(name='Calibri', size=11, bold=True, color='1D4ED8')
                cell_obj.fill = summary_row_fill
                cell_obj.border = grid_border

            ws.row_dimensions[current_row].height = 42

    # ==========================================
    # SUBJECT-WISE REPORT GENERATION
    # ==========================================
    else:
        # Find all schedule sessions for this subject within date range
        schedules_qs = Schedule.objects.all()
        if selected_subject:
            schedules_qs = schedules_qs.filter(subject=selected_subject)
        if selected_class:
            schedules_qs = schedules_qs.filter(student_class=selected_class)

        # Query actual sessions logged in Attendance table
        atts_filter = {}
        if from_date:
            atts_filter['date__gte'] = from_date
        if to_date:
            atts_filter['date__lte'] = to_date

        atts = Attendance.objects.filter(
            student__in=students,
            **atts_filter
        )
        if selected_subject:
            atts = atts.filter(schedule__subject=selected_subject)
        atts = filter_active_attendance(atts).select_related('schedule', 'student')

        # Build list of distinct (date, period, schedule_id) sessions sorted chronologically
        sessions_set = set()
        for a in atts:
            sessions_set.add((a.date, a.schedule.period, a.schedule_id))

        sorted_sessions = sorted(list(sessions_set), key=lambda x: (x[0], x[1]))

        # Base Headers
        headers = ["S.No", "Student Name", "Register Number", "Roll Number", "Class", "Subject"]

        for d_obj, p_num, _ in sorted_sessions:
            headers.append(f"{d_obj.strftime('%d-%b')} (P{p_num})")

        headers.extend(["Total Classes", "Present", "Absent", "OD", "Subject %"])

        # Write Header Row
        for col_idx, text in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=text)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align_center if col_idx > 2 else align_left
            cell.border = grid_border

        ws.row_dimensions[current_row].height = 28
        current_row += 1

        # Map (student_id, date, schedule_id) -> status
        att_map = {(a.student_id, a.date, a.schedule_id): a.status for a in atts}

        class_total_p = 0
        class_total_a = 0
        class_total_od = 0
        class_total_possible_classes = 0

        # Initialize session summaries
        session_p_counts = {session: 0.0 for session in sorted_sessions}
        session_a_counts = {session: 0.0 for session in sorted_sessions}
        session_od_counts = {session: 0.0 for session in sorted_sessions}

        # Populate Student Rows
        for idx, student in enumerate(students, start=1):
            row_num = current_row
            ws.cell(row=row_num, column=1, value=idx).alignment = align_center
            ws.cell(row=row_num, column=2, value=student.user.get_full_name() or student.user.username).alignment = align_left
            ws.cell(row=row_num, column=3, value=student.reg_no or student.user.username).alignment = align_center
            ws.cell(row=row_num, column=4, value=student.roll_no or '-').alignment = align_center
            ws.cell(row=row_num, column=5, value=student.student_class.display_name if student.student_class else '-').alignment = align_center
            ws.cell(row=row_num, column=6, value=selected_subject.code if selected_subject else 'ALL').alignment = align_center

            for c in range(1, 7):
                cell = ws.cell(row=row_num, column=c)
                cell.font = regular_font
                cell.border = grid_border
                if idx % 2 == 0:
                    cell.fill = zebra_fill

            st_p = 0
            st_a = 0
            st_od = 0
            st_verified_od = 0
            st_total_classes = len(sorted_sessions)

            for s_idx, (d_obj, p_num, sched_id) in enumerate(sorted_sessions, start=7):
                status_val = att_map.get((student.user_id, d_obj, sched_id), 'Absent')
                is_verified_od = (student.user_id, d_obj) in verified_od_map

                if status_val == 'Present':
                    status_code = "P"
                    st_p += 1
                elif status_val == 'OD':
                    status_code = "OD"
                    st_od += 1
                    if is_verified_od:
                        st_verified_od += 1
                else:
                    status_code = "A"
                    st_a += 1

                cell = ws.cell(row=row_num, column=s_idx, value=status_code)
                cell.alignment = align_center
                cell.border = grid_border

                if status_code == "P":
                    cell.fill = present_fill
                    cell.font = present_font
                    session_p_counts[(d_obj, p_num, sched_id)] += 1
                elif status_code == "A":
                    cell.fill = absent_fill
                    cell.font = absent_font
                    session_a_counts[(d_obj, p_num, sched_id)] += 1
                elif status_code == "OD":
                    cell.fill = od_fill
                    cell.font = od_font
                    session_od_counts[(d_obj, p_num, sched_id)] += 1

            st_effective_present = st_p + st_verified_od
            att_pct = round((st_effective_present / st_total_classes * 100), 1) if st_total_classes > 0 else 0.0

            sum_col_start = 7 + len(sorted_sessions)
            c_tot = ws.cell(row=row_num, column=sum_col_start, value=st_total_classes)
            c_p = ws.cell(row=row_num, column=sum_col_start+1, value=st_p)
            c_a = ws.cell(row=row_num, column=sum_col_start+2, value=st_a)
            c_od = ws.cell(row=row_num, column=sum_col_start+3, value=st_od)
            c_pct = ws.cell(row=row_num, column=sum_col_start+4, value=f"{att_pct}%")

            for sc_idx, cell_obj in enumerate([c_tot, c_p, c_a, c_od, c_pct], start=sum_col_start):
                cell_obj.alignment = align_center
                cell_obj.font = bold_font
                cell_obj.border = grid_border
                if idx % 2 == 0:
                    cell_obj.fill = zebra_fill

            class_total_p += st_effective_present
            class_total_a += st_a
            class_total_od += st_od
            class_total_possible_classes += st_total_classes

            ws.row_dimensions[row_num].height = 20
            current_row += 1

        # Add Summary Row at Bottom
        if len(students) > 0:
            ws.cell(row=current_row, column=1, value="").border = grid_border
            ws.cell(row=current_row, column=2, value="SUBJECT AVERAGE SUMMARY").font = bold_font
            ws.cell(row=current_row, column=2).alignment = align_left
            ws.cell(row=current_row, column=2).border = grid_border

            for c in range(3, 7):
                cell = ws.cell(row=current_row, column=c, value="-")
                cell.alignment = align_center
                cell.font = bold_font
                cell.border = grid_border
                cell.fill = summary_row_fill

            for s_idx, session in enumerate(sorted_sessions, start=7):
                p_sum = session_p_counts[session]
                a_sum = session_a_counts[session]
                od_sum = session_od_counts[session]

                p_str = int(p_sum) if p_sum.is_integer() else p_sum
                a_str = int(a_sum) if a_sum.is_integer() else a_sum
                od_str = int(od_sum) if od_sum.is_integer() else od_sum

                sum_text = f"P:{p_str}\nA:{a_str}\nO:{od_str}"

                cell = ws.cell(row=current_row, column=s_idx, value=sum_text)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.font = Font(name='Calibri', size=9, bold=True, color='475569')
                cell.border = grid_border
                cell.fill = summary_row_fill

            sum_col_start = 7 + len(sorted_sessions)
            avg_pct = round((class_total_p / class_total_possible_classes * 100), 1) if class_total_possible_classes > 0 else 0.0

            c_tot = ws.cell(row=current_row, column=sum_col_start, value=class_total_possible_classes)
            c_p = ws.cell(row=current_row, column=sum_col_start+1, value=class_total_p)
            c_a = ws.cell(row=current_row, column=sum_col_start+2, value=class_total_a)
            c_od = ws.cell(row=current_row, column=sum_col_start+3, value=class_total_od)
            c_pct = ws.cell(row=current_row, column=sum_col_start+4, value=f"{avg_pct}%")

            for cell_obj in [c_tot, c_p, c_a, c_od, c_pct]:
                cell_obj.alignment = align_center
                cell_obj.font = Font(name='Calibri', size=11, bold=True, color='1D4ED8')
                cell_obj.fill = summary_row_fill
                cell_obj.border = grid_border

            ws.row_dimensions[current_row].height = 42

    # ==========================================
    # AUTO ARRANGE COLUMN WIDTHS ACCORDING TO CONTENT
    # ==========================================
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        
        # Determine maximum content length in this column
        for cell in col:
            # Skip title block rows 1 to 4 for width calculation
            if cell.row < 6:
                continue
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
                
        # Set dynamic column width with padding
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # Construct descriptive filename
    filename_parts = ["Attendance_Report"]
    if report_mode == 'day':
        filename_parts.append("Daily")
    elif report_mode == 'subject_percentage':
        filename_parts.append("Subject_Summary")
    else:
        filename_parts.append(str(report_mode).capitalize())

    if report_type == 'student' and student_id:
        filename_parts.append(f"Student_{student_id}")
    elif report_type == 'tutored':
        filename_parts.append("Tutored_Group")

    if selected_class:
        cls_name = str(selected_class).replace(" - ", "_").replace(" ", "_")
        filename_parts.append(cls_name)

    if selected_subject:
        sub_name = f"{selected_subject.code}_{selected_subject.name}".replace(" ", "_")
        filename_parts.append(sub_name)

    if from_date and to_date:
        if from_date == to_date:
            filename_parts.append(from_date.strftime('%Y-%m-%d'))
        else:
            filename_parts.append(f"{from_date.strftime('%Y-%m-%d')}_to_{to_date.strftime('%Y-%m-%d')}")

    raw_filename = "_".join(filename_parts)
    import re
    clean_filename = re.sub(r'[^a-zA-Z0-9_\-]', '', raw_filename)
    filename = f"{clean_filename}.xlsx"

    return wb, filename


def generate_student_subject_detail_excel(student, subject, records, stats, verified_ods):
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Subject Attendance Log"
    ws.views.sheetView[0].showGridLines = True

    # Styling Palette
    header_fill = PatternFill(start_color='0F172A', end_color='0F172A', fill_type='solid') # Navy
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')

    title_font = Font(name='Calibri', size=14, bold=True, color='0F172A')
    subtitle_font = Font(name='Calibri', size=11, bold=True, color='475569')
    meta_font = Font(name='Calibri', size=10, italic=True, color='64748B')
    bold_font = Font(name='Calibri', size=11, bold=True, color='0F172A')
    regular_font = Font(name='Calibri', size=11, color='1E293B')

    present_fill = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid')
    present_font = Font(name='Calibri', size=11, bold=True, color='15803D')

    absent_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
    absent_font = Font(name='Calibri', size=11, bold=True, color='B91C1C')

    od_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
    od_font = Font(name='Calibri', size=11, bold=True, color='B45309')

    zebra_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
    summary_row_fill = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')

    thin_side = Side(style='thin', color='CBD5E1')
    grid_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')

    # Institutional Header Block
    ws['A1'] = "DR. NGP INSTITUTE OF TECHNOLOGY (AUTONOMOUS)"
    ws['A1'].font = title_font
    ws['A2'] = f"DEPARTMENT OF {(student.student_class.department.name.upper() if student.student_class and student.student_class.department else 'COMPUTER SCIENCE AND ENGINEERING')}"
    ws['A2'].font = subtitle_font
    ws['A3'] = "STUDENT ATTENDANCE DETAILED LOG"
    ws['A3'].font = Font(name='Calibri', size=12, bold=True, color='1D4ED8')

    class_str = student.student_class.display_name if student.student_class else "N/A"
    ws['A4'] = f"Student: {student.user.get_full_name()} ({student.reg_no or student.user.username}) | Class: {class_str} | Subject: {subject.code} - {subject.name}"
    ws['A4'].font = meta_font

    # Stats Row
    ws['A6'] = "TOTAL HOURS"
    ws['B6'] = "PRESENT"
    ws['C6'] = "ABSENT"
    ws['D6'] = "PERCENTAGE"
    for col in ['A', 'B', 'C', 'D']:
        cell = ws[f"{col}6"]
        cell.font = Font(name='Calibri', size=9, bold=True, color='64748B')
        cell.alignment = align_center
        cell.fill = summary_row_fill
        cell.border = grid_border

    ws['A7'] = stats['total_hours']
    ws['B7'] = stats['effective_present']
    ws['C7'] = stats['absent_count']
    ws['D7'] = f"{stats['percentage']}%"
    
    ws['A7'].font = bold_font
    ws['B7'].font = Font(name='Calibri', size=11, bold=True, color='15803D')
    ws['C7'].font = Font(name='Calibri', size=11, bold=True, color='B91C1C')
    ws['D7'].font = Font(name='Calibri', size=11, bold=True, color='1D4ED8' if stats['percentage'] >= 75 else 'B91C1C')
    
    for col in ['A', 'B', 'C', 'D']:
        ws[f"{col}7"].alignment = align_center
        ws[f"{col}7"].border = grid_border

    # Table Headers
    headers = ["S.No", "Date", "Period", "Status", "Note"]
    start_row = 9
    for col_idx, text in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=text)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center if col_idx != 2 else align_left
        cell.border = grid_border
    
    ws.row_dimensions[start_row].height = 24
    
    current_row = 10
    for idx, r in enumerate(records, start=1):
        ws.cell(row=current_row, column=1, value=idx).alignment = align_center
        ws.cell(row=current_row, column=2, value=r.date.strftime('%Y-%m-%d')).alignment = align_left
        ws.cell(row=current_row, column=3, value=f"Period {r.schedule.period}").alignment = align_center
        
        status_cell = ws.cell(row=current_row, column=4, value=r.status)
        status_cell.alignment = align_center
        if r.status == 'Present':
            status_cell.fill = present_fill
            status_cell.font = present_font
        elif r.status == 'Absent':
            status_cell.fill = absent_fill
            status_cell.font = absent_font
        elif r.status == 'OD':
            status_cell.fill = od_fill
            status_cell.font = od_font
            
        note = 'Ignored (Optional 8th Period)' if (r.schedule.period == 8 and r.status != 'Present') else '-'
        ws.cell(row=current_row, column=5, value=note).alignment = align_left
        
        for c in range(1, 6):
            cell = ws.cell(row=current_row, column=c)
            cell.border = grid_border
            if c != 4:
                cell.font = regular_font
            if idx % 2 == 0:
                cell.fill = zebra_fill
                
        ws.row_dimensions[current_row].height = 20
        current_row += 1

    # Auto Column Widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row < 9:
                continue
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    filename = f"Attendance_Log_{student.reg_no or student.user.username}_{subject.code}.xlsx"
    return wb, filename


def generate_subject_attendance_excel(records):
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Subject Attendance"
    ws.views.sheetView[0].showGridLines = True

    # Styling Palette
    header_fill = PatternFill(start_color='0F172A', end_color='0F172A', fill_type='solid') # Navy
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')

    title_font = Font(name='Calibri', size=14, bold=True, color='0F172A')
    subtitle_font = Font(name='Calibri', size=11, bold=True, color='475569')
    meta_font = Font(name='Calibri', size=10, italic=True, color='64748B')
    bold_font = Font(name='Calibri', size=11, bold=True, color='0F172A')
    regular_font = Font(name='Calibri', size=11, color='1E293B')

    present_fill = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid')
    present_font = Font(name='Calibri', size=11, bold=True, color='15803D')

    absent_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
    absent_font = Font(name='Calibri', size=11, bold=True, color='B91C1C')

    od_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
    od_font = Font(name='Calibri', size=11, bold=True, color='B45309')

    zebra_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
    summary_row_fill = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')

    thin_side = Side(style='thin', color='CBD5E1')
    grid_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')

    # Get metadata from first record if exists
    first = records.first() if hasattr(records, 'first') and records.exists() else None
    dept_str = first.student.student_class.department.name.upper() if first and first.student.student_class and first.student.student_class.department else 'ENGINEERING'
    
    ws['A1'] = "DR. NGP INSTITUTE OF TECHNOLOGY (AUTONOMOUS)"
    ws['A1'].font = title_font
    ws['A2'] = f"DEPARTMENT OF {dept_str}"
    ws['A2'].font = subtitle_font
    ws['A3'] = "SUBJECT-WISE ATTENDANCE LOG"
    ws['A3'].font = Font(name='Calibri', size=12, bold=True, color='1D4ED8')
    
    c_name = first.student.student_class.name if first and first.student.student_class else 'N/A'
    s_code = first.schedule.subject.code if first and first.schedule else 'N/A'
    s_name = first.schedule.subject.name if first and first.schedule else 'N/A'
    date_str = first.date.strftime('%Y-%m-%d') if first else 'N/A'
    period_str = str(first.schedule.period) if first and first.schedule else 'N/A'
    ws['A4'] = f"Class: {c_name} | Subject: {s_code} - {s_name} | Period: {period_str} | Date: {date_str}"
    ws['A4'].font = meta_font

    headers = ['S.No', 'Register Number', 'Name', 'Department', 'Year', 'Class', 'Section', 'Date', 'Time', 'Subject', 'Status']
    start_row = 6
    for col_idx, text in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=text)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center if col_idx != 3 else align_left
        cell.border = grid_border

    ws.row_dimensions[start_row].height = 24
    
    current_row = 7
    total_c = 0
    present_c = 0
    absent_c = 0
    od_c = 0
    
    for idx, r in enumerate(records, start=1):
        reg_no = r.student.reg_no or r.student.roll_no or r.student.user.username
        full_name = f"{r.student.user.first_name} {r.student.user.last_name}".strip() or r.student.user.username
        time_str = r.schedule.start_time.strftime('%H:%M') if r.schedule.start_time else ''
        subject_str = r.schedule.subject.name
        dept = r.student.student_class.department.name if r.student.student_class and r.student.student_class.department else ''
        yr = r.student.student_class.year if r.student.student_class else ''
        cls = r.student.student_class.name if r.student.student_class else ''
        sec = r.student.student_class.section if r.student.student_class else ''
        
        ws.cell(row=current_row, column=1, value=idx).alignment = align_center
        ws.cell(row=current_row, column=2, value=reg_no).alignment = align_center
        ws.cell(row=current_row, column=3, value=full_name).alignment = align_left
        ws.cell(row=current_row, column=4, value=dept).alignment = align_center
        ws.cell(row=current_row, column=5, value=yr).alignment = align_center
        ws.cell(row=current_row, column=6, value=cls).alignment = align_center
        ws.cell(row=current_row, column=7, value=sec).alignment = align_center
        ws.cell(row=current_row, column=8, value=r.date.strftime('%Y-%m-%d')).alignment = align_center
        ws.cell(row=current_row, column=9, value=time_str).alignment = align_center
        ws.cell(row=current_row, column=10, value=subject_str).alignment = align_center
        
        status_cell = ws.cell(row=current_row, column=11, value=r.status)
        status_cell.alignment = align_center
        if r.status == 'Present':
            status_cell.fill = present_fill
            status_cell.font = present_font
            present_c += 1
        elif r.status == 'Absent':
            status_cell.fill = absent_fill
            status_cell.font = absent_font
            absent_c += 1
        elif r.status == 'OD':
            status_cell.fill = od_fill
            status_cell.font = od_font
            od_c += 1
            
        total_c += 1
        
        for c in range(1, 12):
            cell = ws.cell(row=current_row, column=c)
            cell.border = grid_border
            if c != 11:
                cell.font = regular_font
            if idx % 2 == 0:
                cell.fill = zebra_fill
                
        ws.row_dimensions[current_row].height = 20
        current_row += 1

    # Summary Row
    ws.cell(row=current_row, column=1, value="").border = grid_border
    ws.cell(row=current_row, column=2, value="SUMMARY").font = bold_font
    ws.cell(row=current_row, column=2).alignment = align_left
    ws.cell(row=current_row, column=2).border = grid_border
    
    for c in range(3, 8):
        ws.cell(row=current_row, column=c, value="-").border = grid_border
        ws.cell(row=current_row, column=c).fill = summary_row_fill
        ws.cell(row=current_row, column=c).alignment = align_center
        
    ws.cell(row=current_row, column=8, value=f"Total: {total_c}").font = bold_font
    ws.cell(row=current_row, column=8).border = grid_border
    ws.cell(row=current_row, column=8).alignment = align_center
    
    ws.cell(row=current_row, column=9, value=f"P: {present_c}").font = Font(name='Calibri', size=11, bold=True, color='15803D')
    ws.cell(row=current_row, column=9).border = grid_border
    ws.cell(row=current_row, column=9).alignment = align_center
    
    ws.cell(row=current_row, column=10, value=f"A: {absent_c}").font = Font(name='Calibri', size=11, bold=True, color='B91C1C')
    ws.cell(row=current_row, column=10).border = grid_border
    ws.cell(row=current_row, column=10).alignment = align_center
    
    ws.cell(row=current_row, column=11, value=f"O: {od_c}").font = Font(name='Calibri', size=11, bold=True, color='B45309')
    ws.cell(row=current_row, column=11).border = grid_border
    ws.cell(row=current_row, column=11).alignment = align_center
    
    ws.row_dimensions[current_row].height = 24
    
    # Auto column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row < 6:
                continue
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    filename = f"Subject_Attendance_{c_name}_{s_code}_Period_{period_str}_{date_str}.xlsx"
    return wb, filename


def generate_advisor_daily_excel(class_obj, date_obj, records):
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Advisor Daily Attendance"
    ws.views.sheetView[0].showGridLines = True

    # Styling Palette
    header_fill = PatternFill(start_color='0F172A', end_color='0F172A', fill_type='solid') # Navy
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')

    title_font = Font(name='Calibri', size=14, bold=True, color='0F172A')
    subtitle_font = Font(name='Calibri', size=11, bold=True, color='475569')
    meta_font = Font(name='Calibri', size=10, italic=True, color='64748B')
    bold_font = Font(name='Calibri', size=11, bold=True, color='0F172A')
    regular_font = Font(name='Calibri', size=11, color='1E293B')

    present_fill = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid')
    present_font = Font(name='Calibri', size=11, bold=True, color='15803D')

    absent_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
    absent_font = Font(name='Calibri', size=11, bold=True, color='B91C1C')

    od_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
    od_font = Font(name='Calibri', size=11, bold=True, color='B45309')

    zebra_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
    summary_row_fill = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')

    thin_side = Side(style='thin', color='CBD5E1')
    grid_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')

    # Headers
    dept_str = class_obj.department.name.upper() if class_obj and class_obj.department else 'ENGINEERING'
    ws['A1'] = "DR. NGP INSTITUTE OF TECHNOLOGY (AUTONOMOUS)"
    ws['A1'].font = title_font
    ws['A2'] = f"DEPARTMENT OF {dept_str}"
    ws['A2'].font = subtitle_font
    ws['A3'] = "ADVISOR DAILY ATTENDANCE LOG"
    ws['A3'].font = Font(name='Calibri', size=12, bold=True, color='1D4ED8')
    ws['A4'] = f"Class: {class_obj.name} (Sec {class_obj.section}) | Date: {date_obj.strftime('%d-%b-%Y')}"
    ws['A4'].font = meta_font

    headers = ['S.No', 'Register Number', 'Name', 'Department', 'Year', 'Class', 'Section']
    for p in range(1, 9):
        headers.append(f"Period {p}")
        
    start_row = 6
    for col_idx, text in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=text)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center if col_idx != 3 else align_left
        cell.border = grid_border

    ws.row_dimensions[start_row].height = 24
    
    # Process student rows
    student_data = {}
    for r in records:
        username = r.student.user.username
        if username not in student_data:
            student_data[username] = {
                'reg_no': r.student.reg_no or r.student.roll_no or username,
                'name': f"{r.student.user.first_name} {r.student.user.last_name}".strip() or username,
                'dept': r.student.student_class.department.name if r.student.student_class and r.student.student_class.department else '',
                'year': r.student.student_class.year if r.student.student_class else '',
                'cls': r.student.student_class.name if r.student.student_class else '',
                'sec': r.student.student_class.section if r.student.student_class else '',
                'periods': {str(p): '-' for p in range(1, 9)}
            }
        if r.schedule:
            student_data[username]['periods'][str(r.schedule.period)] = r.status

    current_row = 7
    sorted_students = sorted(student_data.values(), key=lambda x: x['reg_no'])
    
    for idx, st in enumerate(sorted_students, start=1):
        ws.cell(row=current_row, column=1, value=idx).alignment = align_center
        ws.cell(row=current_row, column=2, value=st['reg_no']).alignment = align_center
        ws.cell(row=current_row, column=3, value=st['name']).alignment = align_left
        ws.cell(row=current_row, column=4, value=st['dept']).alignment = align_center
        ws.cell(row=current_row, column=5, value=st['year']).alignment = align_center
        ws.cell(row=current_row, column=6, value=st['cls']).alignment = align_center
        ws.cell(row=current_row, column=7, value=st['sec']).alignment = align_center
        
        for p in range(1, 9):
            p_val = st['periods'][str(p)]
            cell = ws.cell(row=current_row, column=7 + p, value=p_val)
            cell.alignment = align_center
            if p_val == 'Present':
                cell.fill = present_fill
                cell.font = present_font
            elif p_val == 'Absent':
                cell.fill = absent_fill
                cell.font = absent_font
            elif p_val == 'OD':
                cell.fill = od_fill
                cell.font = od_font
                
        for c in range(1, 16):
            cell = ws.cell(row=current_row, column=c)
            cell.border = grid_border
            if c < 8 or cell.value not in ['Present', 'Absent', 'OD']:
                cell.font = regular_font
            if idx % 2 == 0:
                cell.fill = zebra_fill
                
        ws.row_dimensions[current_row].height = 20
        current_row += 1

    # Auto column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row < 6:
                continue
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    filename = f"Advisor_Daily_Report_{class_obj.name}_{date_obj.strftime('%Y-%m-%d')}.xlsx"
    return wb, filename


def generate_hod_1st_period_excel(target_class, date_obj, matrix):
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "1st Period Attendance"
    ws.views.sheetView[0].showGridLines = True

    # Styling Palette
    header_fill = PatternFill(start_color='0F172A', end_color='0F172A', fill_type='solid') # Navy
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')

    title_font = Font(name='Calibri', size=14, bold=True, color='0F172A')
    subtitle_font = Font(name='Calibri', size=11, bold=True, color='475569')
    meta_font = Font(name='Calibri', size=10, italic=True, color='64748B')
    bold_font = Font(name='Calibri', size=11, bold=True, color='0F172A')
    regular_font = Font(name='Calibri', size=11, color='1E293B')

    present_fill = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid')
    present_font = Font(name='Calibri', size=11, bold=True, color='15803D')

    absent_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
    absent_font = Font(name='Calibri', size=11, bold=True, color='B91C1C')

    od_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
    od_font = Font(name='Calibri', size=11, bold=True, color='B45309')

    zebra_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
    summary_row_fill = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')

    thin_side = Side(style='thin', color='CBD5E1')
    grid_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')

    # Headers
    dept_str = target_class.department.name.upper() if target_class and target_class.department else 'ENGINEERING'
    ws['A1'] = "DR. NGP INSTITUTE OF TECHNOLOGY (AUTONOMOUS)"
    ws['A1'].font = title_font
    ws['A2'] = f"DEPARTMENT OF {dept_str}"
    ws['A2'].font = subtitle_font
    ws['A3'] = "1ST PERIOD MORNING ATTENDANCE REPORT"
    ws['A3'].font = Font(name='Calibri', size=12, bold=True, color='1D4ED8')
    ws['A4'] = f"Class: {target_class.name} (Sec {target_class.section}) | Date: {date_obj.strftime('%d-%b-%Y')}"
    ws['A4'].font = meta_font

    headers = ['S.No', 'Register Number', 'Name', 'Department', 'Year', 'Class', 'Section', 'Date', 'Status']
    start_row = 6
    for col_idx, text in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=text)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center if col_idx != 3 else align_left
        cell.border = grid_border

    ws.row_dimensions[start_row].height = 24
    
    period1_sched = None
    for s in matrix['schedules']:
        if s.period == 1:
            period1_sched = s
            break
            
    current_row = 7
    total = 0
    present = 0
    absent = 0
    od = 0
    
    for idx, row in enumerate(matrix['student_rows'], start=1):
        status_text = '-'
        if period1_sched:
            for st in row['statuses']:
                if st['schedule_id'] == period1_sched.id:
                    status_text = st['status']
                    break
            if status_text == '-':
                status_text = 'Absent'
                
        ws.cell(row=current_row, column=1, value=idx).alignment = align_center
        ws.cell(row=current_row, column=2, value=row['reg_no']).alignment = align_center
        ws.cell(row=current_row, column=3, value=row['name']).alignment = align_left
        ws.cell(row=current_row, column=4, value=target_class.department.name if target_class.department else '').alignment = align_center
        ws.cell(row=current_row, column=5, value=target_class.year).alignment = align_center
        ws.cell(row=current_row, column=6, value=target_class.name).alignment = align_center
        ws.cell(row=current_row, column=7, value=target_class.section).alignment = align_center
        ws.cell(row=current_row, column=8, value=date_obj.strftime('%Y-%m-%d')).alignment = align_center
        
        status_cell = ws.cell(row=current_row, column=9, value=status_text)
        status_cell.alignment = align_center
        if status_text == 'Present':
            status_cell.fill = present_fill
            status_cell.font = present_font
            present += 1
        elif status_text == 'Absent':
            status_cell.fill = absent_fill
            status_cell.font = absent_font
            absent += 1
        elif status_text == 'OD':
            status_cell.fill = od_fill
            status_cell.font = od_font
            od += 1
            
        total += 1
        
        for c in range(1, 10):
            cell = ws.cell(row=current_row, column=c)
            cell.border = grid_border
            if c != 9:
                cell.font = regular_font
            if idx % 2 == 0:
                cell.fill = zebra_fill
                
        ws.row_dimensions[current_row].height = 20
        current_row += 1

    # Summary Row
    ws.cell(row=current_row, column=1, value="").border = grid_border
    ws.cell(row=current_row, column=2, value="SUMMARY").font = bold_font
    ws.cell(row=current_row, column=2).alignment = align_left
    ws.cell(row=current_row, column=2).border = grid_border
    
    for c in range(3, 6):
        ws.cell(row=current_row, column=c, value="-").border = grid_border
        ws.cell(row=current_row, column=c).fill = summary_row_fill
        ws.cell(row=current_row, column=c).alignment = align_center
        
    ws.cell(row=current_row, column=6, value=f"Total: {total}").font = bold_font
    ws.cell(row=current_row, column=6).border = grid_border
    ws.cell(row=current_row, column=6).alignment = align_center
    
    ws.cell(row=current_row, column=7, value=f"P: {present}").font = Font(name='Calibri', size=11, bold=True, color='15803D')
    ws.cell(row=current_row, column=7).border = grid_border
    ws.cell(row=current_row, column=7).alignment = align_center
    
    ws.cell(row=current_row, column=8, value=f"A: {absent}").font = Font(name='Calibri', size=11, bold=True, color='B91C1C')
    ws.cell(row=current_row, column=8).border = grid_border
    ws.cell(row=current_row, column=8).alignment = align_center
    
    ws.cell(row=current_row, column=9, value=f"O: {od}").font = Font(name='Calibri', size=11, bold=True, color='B45309')
    ws.cell(row=current_row, column=9).border = grid_border
    ws.cell(row=current_row, column=9).alignment = align_center
    
    ws.row_dimensions[current_row].height = 24
    
    # Auto column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row < 6:
                continue
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    filename = f"1st_Period_Attendance_{target_class.name.replace(' ', '_')}_{date_obj.strftime('%Y-%m-%d')}.xlsx"
    return wb, filename


def generate_hod_live_grid_excel(target_class, date_obj, matrix):
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Live Attendance Grid"
    ws.views.sheetView[0].showGridLines = True

    # Styling Palette
    header_fill = PatternFill(start_color='0F172A', end_color='0F172A', fill_type='solid') # Navy
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')

    title_font = Font(name='Calibri', size=14, bold=True, color='0F172A')
    subtitle_font = Font(name='Calibri', size=11, bold=True, color='475569')
    meta_font = Font(name='Calibri', size=10, italic=True, color='64748B')
    bold_font = Font(name='Calibri', size=11, bold=True, color='0F172A')
    regular_font = Font(name='Calibri', size=11, color='1E293B')

    present_fill = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid')
    present_font = Font(name='Calibri', size=11, bold=True, color='15803D')

    absent_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
    absent_font = Font(name='Calibri', size=11, bold=True, color='B91C1C')

    od_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
    od_font = Font(name='Calibri', size=11, bold=True, color='B45309')

    zebra_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
    summary_row_fill = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')

    thin_side = Side(style='thin', color='CBD5E1')
    grid_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')

    # Headers
    dept_str = target_class.department.name.upper() if target_class and target_class.department else 'ENGINEERING'
    ws['A1'] = "DR. NGP INSTITUTE OF TECHNOLOGY (AUTONOMOUS)"
    ws['A1'].font = title_font
    ws['A2'] = f"DEPARTMENT OF {dept_str}"
    ws['A2'].font = subtitle_font
    ws['A3'] = "LIVE ATTENDANCE MATRIX GRID"
    ws['A3'].font = Font(name='Calibri', size=12, bold=True, color='1D4ED8')
    ws['A4'] = f"Class: {target_class.name} (Sec {target_class.section}) | Date: {date_obj.strftime('%d-%b-%Y')}"
    ws['A4'].font = meta_font

    headers = ['S.No', 'Register Number', 'Name', 'Department', 'Year', 'Class', 'Section']
    for s in matrix['schedules']:
        headers.append(f"Period {s.period} ({s.subject.code})")
        
    start_row = 6
    for col_idx, text in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=text)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center if col_idx != 3 else align_left
        cell.border = grid_border

    ws.row_dimensions[start_row].height = 24
    
    current_row = 7
    for idx, row in enumerate(matrix['student_rows'], start=1):
        ws.cell(row=current_row, column=1, value=idx).alignment = align_center
        ws.cell(row=current_row, column=2, value=row['reg_no']).alignment = align_center
        ws.cell(row=current_row, column=3, value=row['name']).alignment = align_left
        ws.cell(row=current_row, column=4, value=target_class.department.name if target_class.department else '').alignment = align_center
        ws.cell(row=current_row, column=5, value=target_class.year).alignment = align_center
        ws.cell(row=current_row, column=6, value=target_class.name).alignment = align_center
        ws.cell(row=current_row, column=7, value=target_class.section).alignment = align_center
        
        for p_idx, s in enumerate(matrix['schedules'], start=8):
            status_text = '-'
            for st in row['statuses']:
                if st['schedule_id'] == s.id:
                    status_text = st['status']
                    break
            cell = ws.cell(row=current_row, column=p_idx, value=status_text)
            cell.alignment = align_center
            if status_text == 'Present':
                cell.fill = present_fill
                cell.font = present_font
            elif status_text == 'Absent':
                cell.fill = absent_fill
                cell.font = absent_font
            elif status_text == 'OD':
                cell.fill = od_fill
                cell.font = od_font
                
        num_cols = 7 + len(matrix['schedules'])
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=current_row, column=c)
            cell.border = grid_border
            if c < 8 or cell.value not in ['Present', 'Absent', 'OD']:
                cell.font = regular_font
            if idx % 2 == 0:
                cell.fill = zebra_fill
                
        ws.row_dimensions[current_row].height = 20
        current_row += 1

    # Add Class Summary Row at Bottom
    ws.cell(row=current_row, column=1, value="").border = grid_border
    ws.cell(row=current_row, column=2, value="TOTAL STUDENTS").font = bold_font
    ws.cell(row=current_row, column=2).alignment = align_left
    ws.cell(row=current_row, column=2).border = grid_border
    
    ws.cell(row=current_row, column=3, value=len(matrix['student_rows'])).font = bold_font
    ws.cell(row=current_row, column=3).alignment = align_center
    ws.cell(row=current_row, column=3).border = grid_border
    
    for c in range(4, 8):
        ws.cell(row=current_row, column=c, value="-").border = grid_border
        ws.cell(row=current_row, column=c).fill = summary_row_fill
        ws.cell(row=current_row, column=c).alignment = align_center

    for p_idx, s in enumerate(matrix['schedules'], start=8):
        col_sum = next((cs for cs in matrix['columns_summary'] if cs['schedule_id'] == s.id), None)
        p_val = col_sum['present'] if col_sum else 0
        a_val = col_sum['absent'] if col_sum else 0
        o_val = col_sum['od'] if col_sum else 0
        sum_text = f"P:{p_val}\nA:{a_val}\nO:{o_val}"
        
        cell = ws.cell(row=current_row, column=p_idx, value=sum_text)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.font = Font(name='Calibri', size=9, bold=True, color='475569')
        cell.border = grid_border
        cell.fill = summary_row_fill
        
    ws.row_dimensions[current_row].height = 42

    # Auto column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row < 6:
                continue
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    filename = f"Live_Grid_{target_class.name.replace(' ', '_')}_{date_obj.strftime('%Y-%m-%d')}.xlsx"
    return wb, filename
