import random
import string
import datetime
from math import radians, cos, sin, asin, sqrt
from django.utils import timezone
from django.shortcuts import get_object_or_404
from django.db.models import Count, Q
from rest_framework.decorators import api_view, permission_classes, action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status, viewsets
from datetime import timedelta, date

from .models import OTP, Attendance
from .serializers import OTPSerializer, AttendanceSerializer
from accounts.models import Student, Class, User
from timetable.models import Schedule
from leave.models import Leave

# Helper distance calculator
def haversine(lon1, lat1, lon2, lat2):
    lon1, lat1, lon2, lat2 = map(radians, [lon1, lat1, lon2, lat2])
    dlon = lon2 - lon1 
    dlat = lat2 - lat1 
    a = sin(dlat/2)**2 + cos(lat1) * cos(lat2) * sin(dlon/2)**2
    c = 2 * asin(sqrt(a)) 
    r = 6371000 # Radius of earth in meters
    return c * r

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def api_generate_otp(request):
    if request.user.role not in ['staff', 'hod']:
        return Response({'detail': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)

    department_name = request.data.get('department_name')
    class_name = request.data.get('class_name')
    subject_name = request.data.get('subject_name')
    period_data = request.data.get('period')
    
    if not (class_name and subject_name and period_data):
        return Response({'detail': 'Missing required fields: class_name, subject_name, period'}, status=status.HTTP_400_BAD_REQUEST)
        
    # Resolve period_data into a list of periods
    if isinstance(period_data, list):
        periods = period_data
    elif isinstance(period_data, str):
        if ',' in period_data:
            periods = [p.strip() for p in period_data.split(',') if p.strip()]
        else:
            periods = [period_data]
    else:
        periods = [period_data]

    if not periods:
        return Response({'detail': 'No valid periods selected.'}, status=status.HTTP_400_BAD_REQUEST)

    from accounts.models import Department, Class, Subject
    import random, string

    # 1. Resolve Department
    if department_name:
        dept_str = str(department_name).strip()
        if dept_str.isdigit():
            dept = get_object_or_404(Department, id=int(dept_str))
        else:
            dept, _ = Department.objects.get_or_create(name=dept_str)
    else:
        dept = request.user.department

    # 2. Resolve Class
    class_str = str(class_name).strip()
    if class_str.isdigit():
        student_class = get_object_or_404(Class, id=int(class_str))
    else:
        student_class, _ = Class.objects.get_or_create(
            name=class_str, 
            department=dept,
            defaults={'year': 1, 'section': 'A'}
        )

    # 3. Resolve Subject
    subject_name = subject_name.strip()
    subject = Subject.objects.filter(name=subject_name, department=dept).first()
    if not subject:
        code = ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))
        while Subject.objects.filter(code=code).exists():
            code = ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))
        subject = Subject.objects.create(name=subject_name, code=code, department=dept)

    today = timezone.now().date()
    day_str = today.strftime('%A')

    # Check locks first
    from .models import PeriodLock
    for p in periods:
        p_val = int(p) if str(p).isdigit() else p
        lock = PeriodLock.objects.filter(student_class=student_class, date=today, period=p_val).first()
        if lock and lock.staff != request.user:
            return Response({
                'detail': f'Period {p_val} is already marked/used by {lock.staff.first_name} {lock.staff.last_name} ({lock.staff.username}).'
            }, status=status.HTTP_400_BAD_REQUEST)
    
    # Generate a single 6-digit code shared among all selected periods
    code = ''.join(random.choices(string.digits, k=6))
    
    lat = request.data.get('latitude')
    lng = request.data.get('longitude')
    accuracy = request.data.get('accuracy')
    lat = float(lat) if lat else None
    lng = float(lng) if lng else None
    accuracy = float(accuracy) if accuracy else 10.0

    otps_created = []

    for p in periods:
        p_val = int(p) if str(p).isdigit() else p
        
        # Acquire/Ensure lock exists for current staff
        PeriodLock.objects.get_or_create(
            student_class=student_class,
            date=today,
            period=p_val,
            defaults={'staff': request.user}
        )

        schedule, created = Schedule.objects.get_or_create(
            student_class=student_class,
            subject=subject,
            period=p_val,
            day=day_str,
            defaults={
                'start_time': datetime.time(9, 0),
                'end_time': datetime.time(10, 0)
            }
        )
        
        # Deactivate old OTPs for this schedule today
        OTP.objects.filter(schedule=schedule, is_active=True).update(is_active=False)
        
        otp = OTP.objects.create(code=code, schedule=schedule, staff_latitude=lat, staff_longitude=lng, staff_accuracy=accuracy, creator=request.user)
        otps_created.append(otp)
        
        # Pre-mark all students
        if schedule.subject and schedule.subject.subject_type in ['OPEN_ELECTIVE', 'PROFESSIONAL_ELECTIVE']:
            students = schedule.subject.get_enrolled_students()
        else:
            students = schedule.student_class.get_students()
        for student in students:
            # Check for approved leave/od
            approved_leave = Leave.objects.filter(student=student, date=today, final_status='Approved').first()
            default_status = approved_leave.leave_type if approved_leave else 'Absent'
            
            Attendance.objects.get_or_create(
                student=student, 
                schedule=schedule, 
                date=today,
                defaults={'status': default_status}
            )

            
    if not otps_created:
        return Response({'detail': 'No periods could be resolved.'}, status=status.HTTP_400_BAD_REQUEST)
        
    return Response({
        'detail': 'OTP generated successfully',
        'otp_id': otps_created[0].id,
        'otp_ids': [o.id for o in otps_created],
        'code': code,
        'periods': periods
    })

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def api_verify_otp(request):
    if request.user.role != 'student':
        return Response({'detail': 'Only students can mark attendance'}, status=status.HTTP_403_FORBIDDEN)

    code = request.data.get('otp_code')
    student_lat = request.data.get('latitude')
    student_lng = request.data.get('longitude')
    student_accuracy = request.data.get('accuracy')
    
    if student_lat is None or student_lng is None:
        return Response({'detail': 'Failed to retrieve your location'}, status=status.HTTP_400_BAD_REQUEST)
        
    student_lat = float(student_lat)
    student_lng = float(student_lng)
    
    # Find active OTPs
    otp_qs = OTP.objects.filter(code=code, is_active=True).order_by('-created_at')
    if otp_qs.exists():
        otp = otp_qs.first()
        now = timezone.now()
        
        # Check 3 minute validity (on the most recent one)
        if now <= otp.created_at + timedelta(minutes=3):
            # Geofence check if staff provided location
            if otp.staff_latitude and otp.staff_longitude:
                distance = haversine(student_lng, student_lat, otp.staff_longitude, otp.staff_latitude)
                
                s_acc = float(student_accuracy) if student_accuracy else 10.0
                t_acc = float(otp.staff_accuracy) if otp.staff_accuracy else 10.0
                inaccuracy_buffer = max(0.0, s_acc - 10.0) + max(0.0, t_acc - 10.0)
                allowed_limit = 100.0 + inaccuracy_buffer
                
                if distance > allowed_limit:
                    return Response({
                        'detail': f"You are too far from the teacher's session location (Distance: {distance:.1f}m > limit {allowed_limit:.1f}m)."
                    }, status=status.HTTP_400_BAD_REQUEST)
            
            try:
                student = request.user.student
                today = timezone.localdate()
                
                if otp.schedule.subject and otp.schedule.subject.subject_type in ['OPEN_ELECTIVE', 'PROFESSIONAL_ELECTIVE']:
                    is_enrolled = otp.schedule.subject.elective_students.filter(pk=student.pk).exists()
                else:
                    is_enrolled = (student.student_class == otp.schedule.student_class)

                if not is_enrolled:
                    return Response({'detail': 'You are not enrolled in this class session.'}, status=status.HTTP_400_BAD_REQUEST)
                
                # Check if student has approved Leave/OD today
                from leave.models import Leave
                approved_leave = Leave.objects.filter(student=student, date=today, final_status='Approved').first()
                if approved_leave:
                    return Response({
                        'detail': f'You cannot mark Present because you are approved for {approved_leave.leave_type} today.'
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                marked_periods = []
                for active_otp in otp_qs:
                    # Double check validity of each active otp in the queryset
                    if now > active_otp.created_at + timedelta(minutes=3):
                        active_otp.is_active = False
                        active_otp.save()
                        continue
                        
                    # Check existing attendance record for this schedule
                    existing_att = Attendance.objects.filter(student=student, schedule=active_otp.schedule, date=today).first()
                    if existing_att and existing_att.status in ['Leave', 'OD']:
                        continue
                    
                    # Update attendance to Present
                    attendance, created = Attendance.objects.get_or_create(
                        student=student, 
                        schedule=active_otp.schedule, 
                        date=today,
                        defaults={'status': 'Present'}
                    )
                    if not created:
                        attendance.status = 'Present'
                        attendance.save()
                    marked_periods.append(str(active_otp.schedule.period))
                
                if not marked_periods:
                    return Response({
                        'detail': 'No attendance marked. You might have Leave/OD for the selected period(s).'
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                return Response({'detail': f"Attendance marked as Present successfully for Period(s) {', '.join(marked_periods)}"})
            except Student.DoesNotExist:
                return Response({'detail': 'Student profile not found'}, status=status.HTTP_404_NOT_FOUND)
        else:
            # Mark all as inactive since time is up
            otp_qs.update(is_active=False)
            return Response({'detail': 'OTP has expired'}, status=status.HTTP_400_BAD_REQUEST)
    else:
        return Response({'detail': 'Invalid OTP'}, status=status.HTTP_400_BAD_REQUEST)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def api_stop_session(request):
    if request.user.role not in ['staff', 'hod']:
        return Response({'detail': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
        
    otp_id = request.data.get('otp_id')
    otp_ids = request.data.get('otp_ids', [])
    
    if otp_id:
        OTP.objects.filter(id=otp_id).update(is_active=False)
        
    if otp_ids:
        OTP.objects.filter(id__in=otp_ids).update(is_active=False)
        
    if not otp_id and not otp_ids:
        today = timezone.now().date()
        OTP.objects.filter(creator=request.user, created_at__date=today, is_active=True).update(is_active=False)
        
    return Response({'detail': 'Session stopped successfully'})

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_session_stats(request, otp_id):
    if request.user.role not in ['staff', 'hod']:
        return Response({'detail': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
        
    otp = get_object_or_404(OTP, id=otp_id)
    today = timezone.now().date()
    
    # Students enrolled in this schedule/subject
    if otp.schedule.subject and otp.schedule.subject.subject_type in ['OPEN_ELECTIVE', 'PROFESSIONAL_ELECTIVE']:
        students = otp.schedule.subject.get_enrolled_students()
    else:
        students = Student.objects.filter(student_class=otp.schedule.student_class)

    attendances = Attendance.objects.filter(student__in=students, schedule=otp.schedule, date=today).select_related('student__user', 'student__student_class')
    
    present_count = attendances.filter(status='Present').count()
    absent_count = attendances.filter(status='Absent').count()
    od_count = attendances.filter(status='OD').count()
    leave_count = attendances.filter(status='Leave').count()
    
    all_students_list = [
        {
            'reg_no': a.student.reg_no or a.student.roll_no or a.student.user.username,
            'name': f"{a.student.user.first_name} {a.student.user.last_name}".strip() or a.student.user.username,
            'class_name': str(a.student.student_class) if a.student.student_class else '',
            'status': a.status
        } for a in attendances
    ]
    
    # Calculate time left
    time_elapsed = (timezone.now() - otp.created_at).total_seconds()
    time_left = max(0, 180 - time_elapsed) # 3 minutes validity
    
    if time_left <= 0 and otp.is_active:
        otp.is_active = False
        otp.save()
        
    return Response({
        'present_count': present_count,
        'absent_count': absent_count,
        'od_count': od_count,
        'leave_count': leave_count,
        'all_students': all_students_list,
        'time_left': int(time_left),
        'is_active': otp.is_active and time_left > 0,
        'class_name': otp.schedule.student_class.name,
        'subject_name': otp.schedule.subject.name,
        'period': otp.schedule.period
    })

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_student_stats(request, username):
    student = get_object_or_404(Student, user__username=username)
    
    # Authorize: Only self, tutor, advisor, class advisor, or HOD can see this.
    if request.user.role == 'student' and request.user != student.user:
        return Response({'detail': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
    elif request.user.role == 'staff':
        is_tutor = (request.user == student.tutor)
        is_advisor = (request.user == student.advisor)
        is_class_advisor = (student.student_class and student.student_class.advisor == request.user)
        if not (is_tutor or is_advisor or is_class_advisor):
            return Response({'detail': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
    elif request.user.role == 'hod' and request.user.department != student.student_class.department:
        return Response({'detail': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)

    from timetable.models import Schedule
    from accounts.models import Subject
    if student.student_class:
        non_elective_subjects = Subject.objects.filter(
            student_class=student.student_class
        ).exclude(subject_type__in=['OPEN_ELECTIVE', 'PROFESSIONAL_ELECTIVE'])
        enrolled_elective_subjects = student.elective_subjects.all()
        class_subjects = (non_elective_subjects | enrolled_elective_subjects).distinct()
    else:
        class_subjects = student.elective_subjects.all()

    attendances_qs = Attendance.objects.filter(student=student, schedule__subject__in=class_subjects).select_related('schedule__subject')
    from .models import filter_active_attendance
    attendances_qs = filter_active_attendance(attendances_qs)
    
    is_today = request.query_params.get('today') == 'true'
    if is_today:
        today_date = timezone.localtime(timezone.now()).date()
        attendances_qs = attendances_qs.filter(date=today_date)
        
    attendances = list(attendances_qs)
    total_periods = len(attendances)
    present_periods = sum(1 for a in attendances if a.status == 'Present')
    absent_periods = sum(1 for a in attendances if a.status == 'Absent')
    od_periods = sum(1 for a in attendances if a.status == 'OD')
    leave_periods = sum(1 for a in attendances if a.status == 'Leave')
    
    # Find verified ODs
    from leave.models import Leave
    verified_ods_qs = Leave.objects.filter(
        student=student, 
        leave_type='OD', 
        final_status='Approved', 
        certificate_verified=True
    )
    if is_today:
        verified_ods_qs = verified_ods_qs.filter(date=today_date)
    verified_ods = set(verified_ods_qs.values_list('date', flat=True))
    
    verified_od_count = sum(1 for a in attendances if a.status == 'OD' and a.date in verified_ods)
    
    effective_present = present_periods + verified_od_count
    overall_percentage = (effective_present / total_periods * 100) if total_periods > 0 else 0
    
    # Calculate Days using in-memory aggregation
    att_by_date = {}
    for a in attendances:
        if a.date not in att_by_date:
            att_by_date[a.date] = {'Present': 0, 'OD': 0, 'Absent': 0, 'Leave': 0}
        if a.status in att_by_date[a.date]:
            att_by_date[a.date][a.status] += 1

    dates = list(att_by_date.keys())
    total_days = len(dates)
    present_days = 0
    absent_days = 0
    od_days = 0
    leave_days = 0
    
    for dt, counts in att_by_date.items():
        P = counts['Present']
        O = counts['OD']
        A = counts['Absent']
        L = counts['Leave']
        T = P + O + A + L
        
        is_verified_od_day = (dt in verified_ods)
        if is_verified_od_day:
            verified_od_on_day = O
            unverified_od_on_day = 0
        else:
            verified_od_on_day = 0
            unverified_od_on_day = O
            
        effective_present_on_day = P + verified_od_on_day
        effective_absent_leave_on_day = A + L + unverified_od_on_day
        
        if T > 0:
            if effective_present_on_day >= T / 2.0:
                if verified_od_on_day > P:
                    od_days += 1
                else:
                    present_days += 1
            else:
                if L >= A + unverified_od_on_day:
                    leave_days += 1
                else:
                    absent_days += 1

    # Subject-wise breakdown using in-memory aggregation
    subjects_breakdown = []
    if student.student_class:
        att_by_subject = {}
        for a in attendances:
            sub_id = a.schedule.subject_id
            if sub_id not in att_by_subject:
                att_by_subject[sub_id] = []
            att_by_subject[sub_id].append(a)

        for sub in class_subjects:
            sub_att = att_by_subject.get(sub.id, [])
            sub_total = len(sub_att)
            sub_present = sum(1 for a in sub_att if a.status == 'Present')
            sub_absent = sum(1 for a in sub_att if a.status == 'Absent')
            sub_od = sum(1 for a in sub_att if a.status == 'OD')
            sub_leave = sum(1 for a in sub_att if a.status == 'Leave')
            
            sub_verified_od = sum(1 for a in sub_att if a.status == 'OD' and a.date in verified_ods)
            sub_effective_present = sub_present + sub_verified_od
            sub_percentage = (sub_effective_present / sub_total * 100) if sub_total > 0 else 100.0
            
            subjects_breakdown.append({
                'id': sub.id,
                'name': sub.name,
                'code': sub.code,
                'total_periods': sub_total,
                'present_periods': sub_present,
                'absent_periods': sub_absent,
                'od_periods': sub_od,
                'leave_periods': sub_leave,
                'verified_od_periods': sub_verified_od,
                'effective_present': sub_effective_present,
                'percentage': round(sub_percentage, 2)
            })

    # AI Suggestion
    if overall_percentage >= 90:
        ai_suggestion = f"Excellent! Your attendance is outstanding ({overall_percentage:.2f}%). Keep up the great work to maintain this level of consistency."
    elif overall_percentage >= 75:
        miss_periods = int((4 * effective_present - 3 * total_periods) // 3)
        ai_suggestion = f"Good job! Your attendance is at {overall_percentage:.2f}%, which is above the required 75% threshold."
        if miss_periods > 0:
            ai_suggestion += f" You can afford to miss up to {miss_periods} periods without dropping below 75%."
        else:
            ai_suggestion += " You are close to the limit; try not to miss any more classes."
    else:
        req_periods = int(3 * total_periods - 4 * effective_present)
        ai_suggestion = f"Warning! Your attendance is currently at {overall_percentage:.2f}%, which is below the minimum 75% requirement."
        if req_periods > 0:
            ai_suggestion += f" You need to attend at least {req_periods} consecutive periods without any absence to bring your attendance back to 75%."

    low_subjects = [sub['name'] for sub in subjects_breakdown if sub['percentage'] < 75.0 and sub['total_periods'] > 0]
    if low_subjects:
        ai_suggestion += f" Note: Your attendance in {', '.join(low_subjects)} is below 75%. Prioritize attending these classes."

    return Response({
        'username': student.user.username,
        'name': f"{student.user.first_name} {student.user.last_name}".strip(),
        'profile_photo': student.user.profile_photo or None,
        'class_name': str(student.student_class),
        'total': total_periods,
        'present': present_periods,
        'absent': absent_periods,
        'od': od_periods,
        'verified_od': verified_od_count,
        'percentage': round(overall_percentage, 2),
        'total_periods': total_periods,
        'present_periods': present_periods,
        'absent_periods': absent_periods,
        'od_periods': od_periods,
        'leave_periods': leave_periods,
        'verified_od_periods': verified_od_count,
        'effective_present': effective_present,
        'total_days': total_days,
        'present_days': present_days,
        'absent_days': absent_days,
        'od_days': od_days,
        'leave_days': leave_days,
        'subjects_breakdown': subjects_breakdown,
        'subjects': subjects_breakdown,
        'ai_suggestion': ai_suggestion
    })

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_attendance_report_data(request):
    from_date = request.query_params.get('from_date') or request.query_params.get('date')
    to_date = request.query_params.get('to_date')
    report_type = request.query_params.get('report_type', 'department')
    report_mode = request.query_params.get('report_mode', 'day')
    
    if report_mode == 'day':
        if not to_date:
            to_date = from_date

    if request.user.role == 'student':
        if hasattr(request.user, 'student'):
            records = Attendance.objects.filter(student=request.user.student)
        else:
            records = Attendance.objects.none()
    elif request.user.role == 'hod':
        records = Attendance.objects.filter(student__student_class__department=request.user.department)
    elif request.user.role == 'staff':
        records = Attendance.objects.filter(student__student_class__department=request.user.department)
    elif request.user.role == 'admin':
        records = Attendance.objects.all()
    else:
        records = Attendance.objects.none()

    if report_type == 'class':
        class_id = request.query_params.get('class_id')
        if class_id:
            records = records.filter(student__student_class_id=class_id)
    elif report_type == 'tutored':
        if request.user.role == 'staff':
            records = records.filter(student__tutor=request.user)
    elif report_type == 'student':
        student_id = request.query_params.get('student_id')
        if student_id:
            records = records.filter(Q(student__user__username=student_id) | Q(student__reg_no=student_id))

    if from_date:
        records = records.filter(date__gte=from_date)
    if to_date:
        records = records.filter(date__lte=to_date)

    year = request.query_params.get('year')
    if year:
        try:
            records = records.filter(student__student_class__year=int(year))
        except ValueError:
            pass

    subject_id = request.query_params.get('subject_id')

    if request.user.role == 'staff':
        class_id = request.query_params.get('class_id')
        student_id = request.query_params.get('student_id')
        
        is_related = False
        if report_type == 'class' and class_id:
            try:
                student_class = Class.objects.get(id=class_id)
                is_related = (student_class.advisor == request.user)
            except Class.DoesNotExist:
                pass
        elif report_type == 'student' and student_id:
            try:
                student = Student.objects.get(Q(user__username=student_id) | Q(reg_no=student_id))
                is_related = (student.tutor == request.user or student.advisor == request.user)
            except Student.DoesNotExist:
                pass
        elif report_type == 'tutored':
            is_related = True

        if not is_related and report_mode == 'subject_percentage':
            if not subject_id:
                return Response({'detail': 'Subject is required for this report as you are not the tutor or advisor.'}, status=status.HTTP_400_BAD_REQUEST)
            records = records.filter(schedule__subject_id=subject_id)
        else:
            if subject_id and report_mode == 'subject_percentage':
                records = records.filter(schedule__subject_id=subject_id)
    else:
        if subject_id and report_mode == 'subject_percentage':
            records = records.filter(schedule__subject_id=subject_id)

    result = []
    if report_mode == 'day':
        students_query = Student.objects.all()
        if request.user.role == 'hod' or request.user.role == 'staff':
            students_query = students_query.filter(user__department=request.user.department)
        
        if report_type == 'class' and class_id:
            students_query = students_query.filter(student_class_id=class_id)
        elif report_type == 'tutored':
            students_query = students_query.filter(tutor=request.user)
        elif report_type == 'student' and student_id:
            students_query = students_query.filter(Q(user__username=student_id) | Q(reg_no=student_id))
            
        students_list = list(students_query.select_related('user', 'student_class'))
        
        import datetime
        try:
            start_date = datetime.datetime.strptime(from_date, '%Y-%m-%d').date() if from_date else None
            end_date = datetime.datetime.strptime(to_date, '%Y-%m-%d').date() if to_date else None
        except (ValueError, TypeError):
            start_date = None
            end_date = None

        target_dates = []
        if start_date and end_date:
            if start_date == end_date:
                target_dates = [start_date]
            else:
                db_dates = Attendance.objects.filter(
                    student__in=students_list,
                    date__gte=start_date,
                    date__lte=end_date
                ).values_list('date', flat=True).distinct()
                target_dates = sorted(list(set(db_dates)))
                if not target_dates:
                    target_dates = [start_date]
        else:
            target_dates = [datetime.date.today()]

        records_in_range = Attendance.objects.filter(
            student__in=students_list,
            date__in=target_dates
        ).select_related('student__user', 'student__student_class')
        from .models import filter_active_attendance
        records_in_range = filter_active_attendance(records_in_range)
        
        from collections import defaultdict
        student_date_statuses = defaultdict(list)
        for r in records_in_range:
            student_date_statuses[(r.student_id, r.date)].append(r.status)
            
        attendance_map = {}
        for key, statuses in student_date_statuses.items():
            has_present = 'Present' in statuses or 'OD' in statuses
            has_absent_or_leave = 'Absent' in statuses or 'Leave' in statuses
            if has_present and has_absent_or_leave:
                attendance_map[key] = 'Half Day'
            elif 'OD' in statuses and not has_absent_or_leave:
                attendance_map[key] = 'OD'
            elif 'Leave' in statuses and not has_present:
                attendance_map[key] = 'Leave'
            elif 'Absent' in statuses and not has_present:
                attendance_map[key] = 'Absent'
            else:
                attendance_map[key] = 'Present' if 'Present' in statuses else 'OD'
                
        for d in target_dates:
            date_str = d.strftime('%Y-%m-%d')
            for student in students_list:
                status = attendance_map.get((student.user_id, d), 'Absent')
                result.append({
                    'student_username': student.user.username,
                    'student_reg_no': student.reg_no or student.roll_no or student.user.username,
                    'student_name': f"{student.user.first_name} {student.user.last_name}".strip() or student.user.username,
                    'department_name': student.student_class.department.name if student.student_class and student.student_class.department else '',
                    'year': student.student_class.year if student.student_class else '',
                    'class_name': str(student.student_class),
                    'class_only_name': student.student_class.name if student.student_class else '',
                    'section': student.student_class.section if student.student_class else '',
                    'date': date_str,
                    'status': status
                })
    else:
        from leave.models import Leave
        students_query = Student.objects.all()
        if request.user.role == 'hod' or request.user.role == 'staff':
            students_query = students_query.filter(user__department=request.user.department)
        
        if report_type == 'class' and class_id:
            students_query = students_query.filter(student_class_id=class_id)
        elif report_type == 'tutored':
            students_query = students_query.filter(tutor=request.user)
        elif report_type == 'student' and student_id:
            students_query = students_query.filter(Q(user__username=student_id) | Q(reg_no=student_id))
            
        students_list = students_query.select_related('user', 'student_class')
        
        for student in students_list:
            student_atts = Attendance.objects.filter(student=student)
            from .models import filter_active_attendance
            student_atts = filter_active_attendance(student_atts)
            if from_date:
                student_atts = student_atts.filter(date__gte=from_date)
            if to_date:
                student_atts = student_atts.filter(date__lte=to_date)
            
            if subject_id:
                student_atts = student_atts.filter(schedule__subject_id=subject_id)
                try:
                    subject_label = Subject.objects.get(id=subject_id).name
                except Subject.DoesNotExist:
                    subject_label = 'Subject'
            else:
                subject_label = 'Overall'
                
            total_periods = student_atts.count()
            if total_periods > 0:
                present_periods = student_atts.filter(status='Present').count()
                verified_ods = Leave.objects.filter(
                    student=student, 
                    leave_type='OD', 
                    final_status='Approved', 
                    certificate_verified=True
                ).values_list('date', flat=True)
                verified_od_count = student_atts.filter(status='OD', date__in=verified_ods).count()
                effective_present = present_periods + verified_od_count
                percentage = round((effective_present / total_periods * 100), 2)
            else:
                percentage = 100.0
                
            result.append({
                'student_username': student.user.username,
                'student_reg_no': student.reg_no or student.roll_no or student.user.username,
                'student_name': f"{student.user.first_name} {student.user.last_name}".strip() or student.user.username,
                'department_name': student.student_class.department.name if student.student_class and student.student_class.department else '',
                'year': student.student_class.year if student.student_class else '',
                'class_name': str(student.student_class),
                'class_only_name': student.student_class.name if student.student_class else '',
                'section': student.student_class.section if student.student_class else '',
                'subject_name': subject_label,
                'percentage': percentage
            })

    return Response(result)

@api_view(['GET', 'POST'])
def api_export_excel_report(request):
    user = request.user
    if not user or user.is_anonymous:
        token_str = request.query_params.get('token') or request.query_params.get('auth_token')
        if token_str:
            try:
                from rest_framework_simplejwt.authentication import JWTAuthentication
                jwt_auth = JWTAuthentication()
                validated_token = jwt_auth.get_validated_token(token_str)
                user = jwt_auth.get_user(validated_token)
            except Exception:
                pass
    if not user or user.is_anonymous:
        return Response({'detail': 'Authentication credentials were not provided.'}, status=status.HTTP_401_UNAUTHORIZED)
    
    params = request.data if request.method == 'POST' else request.query_params
    report_mode = params.get('report_mode', 'day')
    report_type = params.get('report_type', 'class')
    student_id = params.get('student_id')
    class_id = params.get('class_id')
    subject_id = params.get('subject_id')
    year = params.get('year')
    from_date_str = params.get('from_date') or params.get('date')
    to_date_str = params.get('to_date') or from_date_str

    # Restrict report access exclusively to Advisor, Tutor, HOD, Admin, or Staff assigned to subject
    if user.role == 'staff':
        staff_type = user.staff.staff_type if hasattr(user, 'staff') else 'Normal'
        from accounts.models import Class, Student, Subject
        is_advisor = Class.objects.filter(advisor=user).exists()
        is_tutor = Student.objects.filter(tutor=user).exists()
        is_assigned_staff = subject_id and Subject.objects.filter(id=subject_id, staff=user).exists()
        
        if staff_type not in ['Advisor', 'Tutor'] and not is_advisor and not is_tutor and not is_assigned_staff:
            return Response(
                {'detail': 'Reports access is reserved for Advisors, Tutors, and assigned Subject Teachers.'},
                status=status.HTTP_403_FORBIDDEN
            )
    elif user.role not in ['hod', 'admin']:
        return Response({'detail': 'Access denied.'}, status=status.HTTP_403_FORBIDDEN)

    is_tutor_role = hasattr(user, 'staff') and user.staff.staff_type == 'Tutor'
    tutor_user = user if (user.role == 'staff' and is_tutor_role) else None

    from .excel_reports import generate_attendance_excel_report
    wb, filename = generate_attendance_excel_report(
        report_mode=report_mode,
        class_id=class_id,
        subject_id=subject_id,
        from_date_str=from_date_str,
        to_date_str=to_date_str,
        tutor_user=tutor_user,
        requested_by_user=user,
        report_type=report_type,
        student_id=student_id,
        year=year
    )

    from django.http import HttpResponse
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_morning_status_excel(request):
    if request.user.role != 'hod':
        return Response({'detail': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
        
    from accounts.models import Class, Student
    from attendance.models import Attendance
    import datetime
    from django.db.models import Count
    
    today = datetime.date.today()
    classes = Class.objects.filter(department=request.user.department).select_related('advisor')
    
    student_counts = Student.objects.filter(student_class__in=classes).values('student_class_id').annotate(count=Count('user_id'))
    att_counts = Attendance.objects.filter(
        student__student_class__in=classes, 
        date=today, 
        schedule__period=1
    ).values('student__student_class_id', 'status').annotate(count=Count('id'))
    
    student_count_map = {item['student_class_id']: item['count'] for item in student_counts}
    att_count_map = {}
    for item in att_counts:
        class_id = item['student__student_class_id']
        status_val = item['status']
        count = item['count']
        if class_id not in att_count_map:
            att_count_map[class_id] = {'Present': 0, 'Absent': 0, 'OD': 0, 'Leave': 0, 'Half Day': 0}
        att_count_map[class_id][status_val] = count
        
    morning_data = []
    for c in classes:
        counts = att_count_map.get(c.id, {'Present': 0, 'Absent': 0, 'OD': 0, 'Leave': 0, 'Half Day': 0})
        total_students = student_count_map.get(c.id, 0)
        morning_data.append({
            'class_name': str(c),
            'year': c.year,
            'section': c.section,
            'total_students': total_students,
            'present_count': counts['Present'],
            'absent_count': counts['Absent'],
            'od_count': counts['OD']
        })
        
    # Generate openpyxl workbook
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Morning Attendance Summary"
    ws.views.sheetView[0].showGridLines = True
    
    # Institution Header
    ws.merge_cells('A1:G1')
    ws['A1'] = "DR. NGP INSTITUTE OF TECHNOLOGY (AUTONOMOUS)"
    ws['A1'].font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='0F172A', end_color='0F172A', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('A2:G2')
    ws['A2'] = f"MORNING ATTENDANCE REPORT (1st Period) - {today.strftime('%d-%b-%Y')}"
    ws['A2'].font = Font(name='Calibri', size=11, bold=True, color='0F172A')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Headers
    headers = ['Class Name', 'Year', 'Section', 'Total Students', 'Present', 'Absent', 'On Duty (OD)']
    ws.append([]) # empty row
    ws.append(headers)
    
    header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    
    for col in range(1, 8):
        cell = ws.cell(row=4, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        
    # Data Rows
    for r_idx, row in enumerate(morning_data, start=5):
        ws.append([
            row['class_name'],
            row['year'],
            row['section'],
            row['total_students'],
            row['present_count'],
            row['absent_count'],
            row['od_count']
        ])
        
        row_fill = PatternFill(start_color='F8FAFC' if r_idx % 2 == 0 else 'FFFFFF', end_color='F8FAFC' if r_idx % 2 == 0 else 'FFFFFF', fill_type='solid')
        
        for col in range(1, 8):
            cell = ws.cell(row=r_idx, column=col)
            cell.fill = row_fill
            cell.border = thin_border
            cell.font = Font(name='Calibri', size=10)
            if col in [1, 3]:
                cell.alignment = Alignment(horizontal='left')
            else:
                cell.alignment = Alignment(horizontal='center')
                
            # Status colored formatting for columns 5, 6, 7
            if col == 5 and cell.value > 0: # Present
                cell.fill = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid')
                cell.font = Font(name='Calibri', size=10, color='15803D', bold=True)
            elif col == 6 and cell.value > 0: # Absent
                cell.fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
                cell.font = Font(name='Calibri', size=10, color='B91C1C', bold=True)
            elif col == 7 and cell.value > 0: # OD
                cell.fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
                cell.font = Font(name='Calibri', size=10, color='B45309', bold=True)

    # Auto fit column widths
    for col in ws.columns:
        vals = [str(cell.value or '') for cell in col if cell.value is not None]
        max_len = max(len(v) for v in vals) if vals else 10
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    ws.row_dimensions[1].height = 35
    ws.row_dimensions[2].height = 25
    ws.row_dimensions[4].height = 25
    
    filename = f"Morning_Attendance_Report_{today.strftime('%Y-%m-%d')}.xlsx"
    from django.http import HttpResponse
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

class AttendanceViewSet(viewsets.ModelViewSet):
    queryset = Attendance.objects.all()
    serializer_class = AttendanceSerializer
    permission_classes = [IsAuthenticated]

    @action(detail=False, methods=['get'], url_path='manual-attendance-data')
    def manual_attendance_data(self, request):
        user = self.request.user
        if user.role != 'staff':
            return Response({'detail': 'Only staff members can access manual attendance.'}, status=status.HTTP_403_FORBIDDEN)
            
        student_id = self.request.query_params.get('student_id')
        date_str = self.request.query_params.get('date')
        
        department = user.department
        students = Student.objects.filter(user__department=department).select_related('user', 'student_class').order_by('reg_no', 'user__username')
        
        students_list = [
            {
                'id': s.user_id,
                'username': s.user.username,
                'name': f"{s.user.first_name} {s.user.last_name}".strip(),
                'class_name': s.student_class.name if s.student_class else '',
                'reg_no': s.reg_no or s.roll_no or s.user.username,
            } for s in students
        ]
        
        if not date_str:
            date_str = timezone.localdate().strftime('%Y-%m-%d')
            
        schedules_data = []
        error_message = None
        
        if student_id:
            try:
                selected_student = Student.objects.get(pk=student_id, user__department=department)
                target_date = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()
                weekday = target_date.strftime('%A')
                
                if selected_student.student_class:
                    schedules = list(Schedule.objects.filter(
                        student_class=selected_student.student_class,
                        day=weekday
                    ).select_related('subject').order_by('period'))
                    
                    atts = Attendance.objects.filter(
                        student=selected_student,
                        schedule__in=schedules,
                        date=target_date
                    )
                    att_dict = {a.schedule_id: a.status for a in atts}
                    
                    for sched in schedules:
                        status_val = att_dict.get(sched.id, 'Absent')
                        schedules_data.append({
                            'schedule_id': sched.id,
                            'subject_name': sched.subject.name,
                            'subject_code': sched.subject.code,
                            'period': sched.period,
                            'status': status_val
                        })
                else:
                    error_message = "Selected student has no assigned class."
            except Student.DoesNotExist:
                error_message = "Student not found in your department."
            except ValueError:
                error_message = "Invalid date format."
                
        return Response({
            'students': students_list,
            'selected_student_id': int(student_id) if student_id and student_id.isdigit() else None,
            'selected_date_str': date_str,
            'schedules_data': schedules_data,
            'error_message': error_message
        })

    @action(detail=False, methods=['post'], url_path='save-manual-attendance')
    def save_manual_attendance(self, request):
        user = self.request.user
        if user.role != 'staff':
            return Response({'detail': 'Only staff members can mark manual attendance.'}, status=status.HTTP_403_FORBIDDEN)
            
        student_id = self.request.data.get('student_id')
        date_str = self.request.data.get('date')
        status_updates = self.request.data.get('statuses', {})
        
        if not student_id or not date_str:
            return Response({'detail': 'Missing student_id or date.'}, status=status.HTTP_400_BAD_REQUEST)
            
        try:
            student = Student.objects.get(pk=student_id, user__department=user.department)
            target_date = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()
            weekday = target_date.strftime('%A')
            
            if not student.student_class:
                return Response({'detail': 'Student has no assigned class.'}, status=status.HTTP_400_BAD_REQUEST)
                
            schedules = Schedule.objects.filter(student_class=student.student_class, day=weekday)
            updated_count = 0
            for sched in schedules:
                status_val = status_updates.get(str(sched.id)) or status_updates.get(sched.id)
                if status_val in ['Present', 'Absent', 'OD', 'Leave']:
                    Attendance.objects.update_or_create(
                        student=student,
                        schedule=sched,
                        date=target_date,
                        defaults={'status': status_val}
                    )
                    updated_count += 1
            return Response({'detail': f'Successfully updated attendance for {updated_count} periods.'})
        except Student.DoesNotExist:
            return Response({'detail': 'Student not found in your department.'}, status=status.HTTP_404_NOT_FOUND)
        except ValueError:
            return Response({'detail': 'Invalid date format.'}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({'detail': f'Error saving attendance: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'], url_path='locked-periods')
    def locked_periods(self, request):
        class_id = request.query_params.get('class_id')
        date_str = request.query_params.get('date')
        if not (class_id and date_str):
            return Response({'detail': 'class_id and date are required.'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            target_date = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            return Response({'detail': 'Invalid date format.'}, status=status.HTTP_400_BAD_REQUEST)
            
        from .models import PeriodLock
        locks = PeriodLock.objects.filter(student_class_id=class_id, date=target_date)
        locks_data = [
            {
                'period': l.period,
                'locked_by_id': l.staff_id,
                'locked_by_name': f"{l.staff.first_name} {l.staff.last_name}".strip() or l.staff.username
            } for l in locks
        ]
        return Response(locks_data)

    @action(detail=False, methods=['get'], url_path='subject-detail')
    def subject_detail(self, request):
        student_username = request.query_params.get('student_username')
        subject_id = request.query_params.get('subject_id')
        if not (student_username and subject_id):
            return Response({'detail': 'student_username and subject_id are required.'}, status=status.HTTP_400_BAD_REQUEST)
            
        from django.db.models import Q
        student = Student.objects.filter(
            Q(user__username=student_username) |
            Q(reg_no=student_username) |
            Q(roll_no=student_username)
        ).first()
        if not student and str(student_username).isdigit():
            student = Student.objects.filter(pk=int(student_username)).first()
        if not student:
            return Response({'detail': f'Student {student_username} not found.'}, status=status.HTTP_404_NOT_FOUND)
        records = Attendance.objects.filter(
            student=student,
            schedule__subject_id=subject_id
        ).select_related('schedule', 'schedule__subject').order_by('-date', 'schedule__period')
        
        from .models import filter_active_attendance
        filtered_records = filter_active_attendance(records)
        
        total_hours = filtered_records.count()
        present_count = filtered_records.filter(status='Present').count()
        absent_count = filtered_records.filter(status='Absent').count()
        od_count = filtered_records.filter(status='OD').count()
        leave_count = filtered_records.filter(status='Leave').count()
        
        from leave.models import Leave
        verified_ods = Leave.objects.filter(
            student=student, 
            leave_type='OD', 
            final_status='Approved', 
            certificate_verified=True
        ).values_list('date', flat=True)
        verified_od_count = filtered_records.filter(status='OD', date__in=verified_ods).count()
        effective_present = present_count + verified_od_count
        percentage = (effective_present / total_hours * 100) if total_hours > 0 else 100.0
        
        from accounts.models import Subject
        subject = get_object_or_404(Subject, id=subject_id)
        
        download = request.query_params.get('download') == 'true'
        if download:
            from .excel_reports import generate_student_subject_detail_excel
            stats_dict = {
                'total_hours': total_hours,
                'effective_present': effective_present,
                'absent_count': absent_count,
                'leave_count': leave_count,
                'percentage': round(percentage, 2)
            }
            wb, filename = generate_student_subject_detail_excel(
                student=student,
                subject=subject,
                records=filtered_records,
                stats=stats_dict,
                verified_ods=verified_ods
            )
            from django.http import HttpResponse
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            wb.save(response)
            return response
            
        records_data = [
            {
                'date': r.date.strftime('%Y-%m-%d'),
                'period': r.schedule.period,
                'status': r.status,
                'ignored': r.schedule.period == 8 and r.status != 'Present'
            } for r in records
        ]
        
        return Response({
            'student_details': {
                'name': f"{student.user.first_name} {student.user.last_name}".strip() or student.user.username,
                'username': student.user.username,
                'reg_no': student.reg_no or student.roll_no or student.user.username,
                'class_name': str(student.student_class),
                'department': student.student_class.department.name if student.student_class and student.student_class.department else '',
            },
            'subject_details': {
                'id': subject.id,
                'name': subject.name,
                'code': subject.code,
            },
            'stats': {
                'total_hours': total_hours,
                'present_count': present_count,
                'absent_count': absent_count,
                'od_count': od_count,
                'leave_count': leave_count,
                'verified_od_count': verified_od_count,
                'effective_present': effective_present,
                'percentage': round(percentage, 2),
            },
            'records': records_data
        })

    @action(detail=False, methods=['get'], url_path='advisor-subject-report')
    def advisor_subject_report(self, request):
        user = self.request.user
        if user.role not in ['staff', 'hod']:
            return Response({'detail': 'Only staff and HOD members can access subject reports.'}, status=status.HTTP_403_FORBIDDEN)
            
        subject_id = request.query_params.get('subject_id')
        if not subject_id:
            return Response({'detail': 'subject_id is required.'}, status=status.HTTP_400_BAD_REQUEST)
            
        from accounts.models import Subject, Student, Class
        subject = get_object_or_404(Subject, id=subject_id)
        if subject.subject_type in ['OPEN_ELECTIVE', 'PROFESSIONAL_ELECTIVE']:
            students = subject.get_enrolled_students().select_related('user', 'student_class').order_by('student_class__name', 'student_class__section', 'reg_no', 'user__username')
        else:
            advised_class = Class.objects.filter(advisor=user).first()
            target_class = subject.student_class or advised_class
            if target_class:
                students = target_class.get_students().select_related('user', 'student_class').order_by('student_class__name', 'student_class__section', 'reg_no', 'user__username')
            else:
                students = Student.objects.none()
        
        from timetable.models import Schedule
        schedules = Schedule.objects.filter(subject=subject)
        
        records = Attendance.objects.filter(
            schedule__subject=subject,
            student__in=students
        ).select_related('student__user', 'schedule').order_by('date', 'schedule__period')
        
        date_periods = sorted(list(set((r.date, r.schedule.period) for r in records)))
        
        from .excel_reports import generate_attendance_excel_report
        wb, filename = generate_attendance_excel_report(
            report_mode='subject_percentage',
            class_id=target_class.id if target_class else None,
            subject_id=subject.id,
            tutor_user=None,
            requested_by_user=user,
            report_type='class'
        )
        from django.http import HttpResponse
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    @action(detail=False, methods=['get'], url_path='advisor-subject-report-json')
    def advisor_subject_report_json(self, request):
        user = self.request.user
        if user.role not in ['staff', 'hod']:
            return Response({'detail': 'Only staff and HOD members can access subject details.'}, status=status.HTTP_403_FORBIDDEN)
            
        subject_id = request.query_params.get('subject_id')
        if not subject_id:
            return Response({'detail': 'subject_id is required.'}, status=status.HTTP_400_BAD_REQUEST)
            
        from accounts.models import Subject, Student, Class
        subject = get_object_or_404(Subject, id=subject_id)
        class_id = request.query_params.get('class_id')
        if subject.subject_type in ['OPEN_ELECTIVE', 'PROFESSIONAL_ELECTIVE'] and not class_id:
            students = subject.get_enrolled_students().select_related('user', 'student_class').order_by('student_class__name', 'student_class__section', 'reg_no', 'user__username')
        else:
            if class_id:
                target_class = Class.objects.filter(id=class_id).first()
            else:
                advised_class = Class.objects.filter(advisor=user).first()
                target_class = subject.student_class or advised_class
                
            if target_class:
                students = target_class.get_students().select_related('user', 'student_class').order_by('student_class__name', 'student_class__section', 'reg_no', 'user__username')
                if subject.subject_type in ['OPEN_ELECTIVE', 'PROFESSIONAL_ELECTIVE']:
                    students = students.filter(id__in=subject.elective_students.values_list('user_id', flat=True))
            else:
                students = Student.objects.none()
        
        records = Attendance.objects.filter(
            schedule__subject=subject,
            student__in=students
        ).select_related('student__user', 'schedule')
        
        from .models import filter_active_attendance
        
        student_data = []
        for student in students:
            student_records = records.filter(student=student)
            filtered_student_records = filter_active_attendance(student_records)
            
            total_hours = filtered_student_records.count()
            present_count = filtered_student_records.filter(status='Present').count()
            absent_count = filtered_student_records.filter(status='Absent').count()
            od_count = filtered_student_records.filter(status='OD').count()
            leave_count = filtered_student_records.filter(status='Leave').count()
            
            from leave.models import Leave
            verified_ods = Leave.objects.filter(
                student=student, 
                leave_type='OD', 
                final_status='Approved', 
                certificate_verified=True
            ).values_list('date', flat=True)
            verified_od_count = filtered_student_records.filter(status='OD', date__in=verified_ods).count()
            effective_present = present_count + verified_od_count
            percentage = (effective_present / total_hours * 100) if total_hours > 0 else 100.0
            
            student_data.append({
                'id': student.pk,
                'reg_no': student.reg_no or student.roll_no or student.user.username,
                'name': f"{student.user.first_name} {student.user.last_name}".strip() or student.user.username,
                'class_id': student.student_class.id if student.student_class else None,
                'class_name': str(student.student_class) if student.student_class else 'Unassigned',
                'class_only_name': student.student_class.name if student.student_class else '',
                'section': student.student_class.section if student.student_class else '',
                'total_hours': total_hours,
                'present_count': present_count,
                'absent_count': absent_count,
                'od_count': od_count,
                'leave_count': leave_count,
                'percentage': round(percentage, 2)
            })
            
        return Response({
            'subject_id': subject.id,
            'subject_name': subject.name,
            'subject_code': subject.code,
            'subject_type': subject.subject_type,
            'students': student_data
        })

    @action(detail=False, methods=['get'], url_path='session-download')
    def session_download(self, request):
        user = self.request.user
        if user.role not in ['staff', 'hod']:
            return Response({'detail': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
            
        class_id = request.query_params.get('class_id')
        period = request.query_params.get('period')
        date = request.query_params.get('date')
        
        if not (class_id and period and date):
            return Response({'detail': 'class_id, period, and date are required.'}, status=status.HTTP_400_BAD_REQUEST)
            
        records = Attendance.objects.filter(
            schedule__student_class_id=class_id,
            schedule__period=period,
            date=date
        ).select_related('student__user', 'student__student_class', 'student__student_class__department', 'schedule__subject')
        
        from .excel_reports import generate_subject_attendance_excel
        wb, filename = generate_subject_attendance_excel(records)
        
        from django.http import HttpResponse
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    @action(detail=False, methods=['get'], url_path='advisor-live-download')
    def advisor_live_download(self, request):
        user = self.request.user
        if user.role != 'staff':
            return Response({'detail': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
            
        from accounts.models import Class
        from accounts.utils import get_live_class_attendance_matrix
        advised_class = Class.objects.filter(advisor=user).first()
        if not advised_class:
            return Response({'detail': 'You do not advise any class.'}, status=status.HTTP_404_NOT_FOUND)
            
        matrix = get_live_class_attendance_matrix(advised_class)
        session_date = matrix['date']
        
        action_type = request.query_params.get('action_type', 'grid') # '1st' or 'grid'
        from .excel_reports import generate_hod_1st_period_excel, generate_hod_live_grid_excel
        if action_type == '1st':
            wb, filename = generate_hod_1st_period_excel(advised_class, session_date, matrix)
        else:
            wb, filename = generate_hod_live_grid_excel(advised_class, session_date, matrix)
            
        from django.http import HttpResponse
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    @action(detail=False, methods=['get'], url_path='hod-morning-download')
    def hod_morning_download(self, request):
        user = self.request.user
        if user.role != 'hod':
            return Response({'detail': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
            
        class_id = request.query_params.get('class_id')
        if not class_id:
            return Response({'detail': 'class_id is required.'}, status=status.HTTP_400_BAD_REQUEST)
            
        from accounts.models import Class
        from accounts.utils import get_live_class_attendance_matrix
        target_class = get_object_or_404(Class, id=class_id, department=user.department)
        matrix = get_live_class_attendance_matrix(target_class)
        session_date = matrix['date']
        
        action_type = request.query_params.get('action_type', 'grid') # '1st' or 'grid'
        from .excel_reports import generate_hod_1st_period_excel, generate_hod_live_grid_excel
        if action_type == '1st':
            wb, filename = generate_hod_1st_period_excel(target_class, session_date, matrix)
        else:
            wb, filename = generate_hod_live_grid_excel(target_class, session_date, matrix)
            
        from django.http import HttpResponse
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    @action(detail=False, methods=['get'], url_path='manual-class-students')
    def manual_class_students(self, request):
        user = self.request.user
        if user.role not in ['staff', 'hod']:
            return Response({'detail': 'Only staff and HOD members can access manual attendance.'}, status=status.HTTP_403_FORBIDDEN)

        class_id = request.query_params.get('class_id')
        subject_id = request.query_params.get('subject_id')
        date_str = request.query_params.get('date')
        period = request.query_params.get('period')

        if not (class_id and subject_id and date_str):
            return Response({'detail': 'Missing class_id, subject_id, or date.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            target_date = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            return Response({'detail': 'Invalid date format. Use YYYY-MM-DD.'}, status=status.HTTP_400_BAD_REQUEST)

        from accounts.models import Subject
        subject = Subject.objects.filter(id=subject_id).first()
        if subject and subject.subject_type in ['OPEN_ELECTIVE', 'PROFESSIONAL_ELECTIVE']:
            students = subject.get_enrolled_students().select_related('user', 'student_class').order_by('student_class__name', 'student_class__section', 'reg_no', 'user__username')
        else:
            students = Student.objects.filter(student_class_id=class_id).select_related('user').order_by('reg_no', 'user__username')

        # Retrieve existing attendance for this class, subject, and date
        weekday = target_date.strftime('%A')
        
        schedules_filter = {
            'subject_id': subject_id,
            'day': weekday
        }
        if not (subject and subject.subject_type in ['OPEN_ELECTIVE', 'PROFESSIONAL_ELECTIVE']):
            schedules_filter['student_class_id'] = class_id

        period_val = None
        if period:
            try:
                period_val = int(period)
                schedules_filter['period'] = period_val
            except ValueError:
                pass

        schedules = Schedule.objects.filter(**schedules_filter)
        
        existing_attendance = {}
        if schedules.exists():
            attendances = Attendance.objects.filter(
                student__in=students,
                schedule__in=schedules,
                date=target_date
            )
            for att in attendances:
                # If there are multiple periods, just pick the status of the first one we find
                existing_attendance[att.student_id] = att.status

        # Format students list
        students_list = [
            {
                'id': s.user_id,
                'username': s.user.username,
                'name': f"{s.user.first_name} {s.user.last_name}".strip() or s.user.username,
                'class_name': str(s.student_class) if s.student_class else '',
                'reg_no': s.reg_no or s.roll_no or s.user.username,
                'roll_no': s.roll_no or '',
                'current_status': existing_attendance.get(s.user_id, 'Present')  # Default to Present if not marked
            } for s in students
        ]

        # Check if weekly schedule exists for this class, subject on this weekday
        schedule_exists = schedules.exists()

        # Check if this period is locked
        from .models import PeriodLock
        is_locked = False
        locked_by_name = ""
        if period_val:
            lock = PeriodLock.objects.filter(student_class_id=class_id, date=target_date, period=period_val).first()
            if lock:
                from accounts.models import Class
                student_class = Class.objects.filter(id=class_id).first()
                is_advisor = student_class and (student_class.advisor == user)
                
                if is_advisor:
                    is_locked = True
                    locked_by_name = "Advisor (must edit through HOD/Advisor Whole Day Attendance)"
                else:
                    is_locked = (lock.staff != user)
                    locked_by_name = f"{lock.staff.first_name} {lock.staff.last_name}".strip() or lock.staff.username

        return Response({
            'students': students_list,
            'schedule_exists': schedule_exists,
            'weekday': weekday,
            'is_locked': is_locked,
            'locked_by_name': locked_by_name
        })

    @action(detail=False, methods=['get'], url_path='class-period-locks')
    def class_period_locks(self, request):
        class_id = request.query_params.get('class_id')
        date_str = request.query_params.get('date')
        if not (class_id and date_str):
            return Response({'locks': []})

        try:
            target_date = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            return Response({'locks': []})

        from .models import PeriodLock
        locks = PeriodLock.objects.filter(student_class_id=class_id, date=target_date)
        res = []
        for l in locks:
            res.append({
                'period': l.period,
                'locked_by_id': l.staff_id,
                'locked_by_name': f"{l.staff.first_name} {l.staff.last_name}".strip() or l.staff.username
            })
        return Response({'locks': res})

    @action(detail=False, methods=['post'], url_path='save-class-manual-attendance')
    def save_class_manual_attendance(self, request):
        user = self.request.user
        if user.role not in ['staff', 'hod']:
            return Response({'detail': 'Only staff and HOD members can access manual attendance.'}, status=status.HTTP_403_FORBIDDEN)

        class_id = request.data.get('class_id')
        subject_id = request.data.get('subject_id')
        date_str = request.data.get('date')
        period = request.data.get('period')
        periods_input = request.data.get('periods')
        statuses = request.data.get('statuses', {})

        if not (class_id and subject_id and date_str):
            return Response({'detail': 'Missing class_id, subject_id, or date.'}, status=status.HTTP_400_BAD_REQUEST)

        periods_list = []
        if periods_input and isinstance(periods_input, list):
            for p in periods_input:
                if str(p).isdigit():
                    periods_list.append(int(p))
        elif period is not None and str(period).isdigit():
            periods_list.append(int(period))

        if not periods_list:
            periods_list = [1]

        try:
            target_date = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()
            weekday = target_date.strftime('%A')
        except ValueError:
            return Response({'detail': 'Invalid date format.'}, status=status.HTTP_400_BAD_REQUEST)

        from .models import PeriodLock
        from accounts.models import Class, Subject

        student_class = get_object_or_404(Class, id=class_id)
        subject = get_object_or_404(Subject, id=subject_id)
        is_advisor = (student_class.advisor == user)

        # Validate PeriodLock for all selected periods
        for p_val in periods_list:
            lock = PeriodLock.objects.filter(student_class=student_class, date=target_date, period=p_val).first()
            if lock:
                if is_advisor:
                    return Response({
                        'detail': f'Period {p_val} is marked. As advisor, please edit through Advisor Whole Day Manual Attendance.'
                    }, status=status.HTTP_400_BAD_REQUEST)
                elif lock.staff != user:
                    return Response({
                        'detail': f'Period {p_val} attendance is already marked by {lock.staff.first_name} {lock.staff.last_name} ({lock.staff.username}).'
                    }, status=status.HTTP_400_BAD_REQUEST)

        from django.db import transaction
        try:
            with transaction.atomic():
                schedules_to_update = []
                for p_val in periods_list:
                    PeriodLock.objects.get_or_create(
                        student_class=student_class,
                        date=target_date,
                        period=p_val,
                        defaults={'staff': user}
                    )

                    sched = Schedule.objects.filter(student_class=student_class, subject=subject, day=weekday, period=p_val).first()
                    if not sched:
                        start_hour = 9 + (p_val - 1)
                        if p_val >= 5:
                            start_hour += 1
                        start_time = datetime.time(start_hour, 0)
                        end_time = datetime.time(start_hour + 1, 0)

                        sched = Schedule.objects.create(
                            student_class=student_class,
                            subject=subject,
                            period=p_val,
                            day=weekday,
                            start_time=start_time,
                            end_time=end_time
                        )
                    schedules_to_update.append(sched)

                students = Student.objects.filter(student_class=student_class)
                updated_count = 0
                for student in students:
                    status_val = statuses.get(str(student.user_id)) or statuses.get(student.user_id) or 'Present'
                    if status_val not in ['Present', 'Absent', 'OD']:
                        status_val = 'Present'

                    for s_item in schedules_to_update:
                        Attendance.objects.update_or_create(
                            student=student,
                            schedule=s_item,
                            date=target_date,
                            defaults={'status': status_val}
                        )
                        updated_count += 1

                p_str = ", ".join([f"P{p}" for p in sorted(periods_list)])
                return Response({
                    'success': True,
                    'detail': f'Successfully marked attendance for {students.count()} students across {p_str} ({updated_count} slot records).'
                })
        except Exception as e:
            return Response({'detail': f'Error saving attendance: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'], url_path='advisor-class-students')
    def advisor_class_students(self, request):
        user = self.request.user
        if user.role not in ['staff', 'hod']:
            return Response({'detail': 'Only staff and HOD members can access advisor manual attendance.'}, status=status.HTTP_403_FORBIDDEN)
        
        from accounts.models import Class
        advised_class = Class.objects.filter(advisor=user).first()
        is_advisor = (hasattr(user, 'staff') and user.staff.staff_type == 'Advisor') or advised_class is not None
        if not is_advisor:
            return Response({'detail': 'Only Advisors can access advisor manual attendance.'}, status=status.HTTP_403_FORBIDDEN)
            
        if not advised_class:
            return Response({'detail': 'You are not assigned as an advisor to any class.'}, status=status.HTTP_400_BAD_REQUEST)
            
        date_str = request.query_params.get('date')
        if not date_str:
            date_str = timezone.localdate().strftime('%Y-%m-%d')
            
        try:
            target_date = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()
            weekday = target_date.strftime('%A')
        except ValueError:
            return Response({'detail': 'Invalid date format. Use YYYY-MM-DD.'}, status=status.HTTP_400_BAD_REQUEST)
            
        students = Student.objects.filter(student_class=advised_class).select_related('user').order_by('reg_no', 'user__username')
        
        # Get schedules for this class on this weekday
        schedules = Schedule.objects.filter(student_class=advised_class, day=weekday).order_by('period')
        
        # We want to represent 8 periods (1 to 8)
        periods_list = []
        for period in range(1, 9):
            sched = schedules.filter(period=period).first()
            periods_list.append({
                'period': period,
                'subject_name': sched.subject.name if sched else 'No Schedule',
                'subject_code': sched.subject.code if sched else '',
                'schedule_id': sched.id if sched else None
            })
            
        # Get existing attendance for this class and date
        existing_attendances = Attendance.objects.filter(
            student__student_class=advised_class,
            date=target_date
        ).select_related('schedule')
        
        # Map student_id -> period -> status
        att_map = {}
        for att in existing_attendances:
            s_id = att.student_id
            p_num = att.schedule.period
            if s_id not in att_map:
                att_map[s_id] = {}
            att_map[s_id][p_num] = att.status
            
        students_data = []
        for s in students:
            # By default, all periods are 'Present' unless already marked in DB
            statuses = {}
            for p in range(1, 9):
                statuses[str(p)] = att_map.get(s.user_id, {}).get(p, 'Present')
                
            students_data.append({
                'id': s.user_id,
                'username': s.user.username,
                'name': f"{s.user.first_name} {s.user.last_name}".strip() or s.user.username,
                'reg_no': s.reg_no or s.roll_no or s.user.username,
                'roll_no': s.roll_no or '',
                'statuses': statuses
            })
            
        return Response({
            'class_id': advised_class.id,
            'class_name': str(advised_class),
            'date': date_str,
            'weekday': weekday,
            'periods': periods_list,
            'students': students_data
        })

    @action(detail=False, methods=['post'], url_path='save-advisor-manual-attendance')
    def save_advisor_manual_attendance(self, request):
        user = self.request.user
        if user.role not in ['staff', 'hod']:
            return Response({'detail': 'Only staff and HOD members can mark manual attendance.'}, status=status.HTTP_403_FORBIDDEN)
            
        from accounts.models import Class
        advised_class = Class.objects.filter(advisor=user).first()
        is_advisor = (hasattr(user, 'staff') and user.staff.staff_type == 'Advisor') or advised_class is not None
        if not is_advisor:
            return Response({'detail': 'Only Advisors can mark advisor manual attendance.'}, status=status.HTTP_403_FORBIDDEN)
            
        if not advised_class:
            return Response({'detail': 'You are not assigned as an advisor to any class.'}, status=status.HTTP_400_BAD_REQUEST)
            
        date_str = request.data.get('date')
        attendance_data = request.data.get('attendance_data', {})
        
        if not date_str:
            return Response({'detail': 'Missing date.'}, status=status.HTTP_400_BAD_REQUEST)
            
        try:
            target_date = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()
            weekday = target_date.strftime('%A')
        except ValueError:
            return Response({'detail': 'Invalid date format.'}, status=status.HTTP_400_BAD_REQUEST)

        from django.db import transaction
        try:
            with transaction.atomic():
                # Acquire/Update locks for periods 1 to 8
                from .models import PeriodLock
                for p in range(1, 9):
                    PeriodLock.objects.update_or_create(
                        student_class=advised_class,
                        date=target_date,
                        period=p,
                        defaults={'staff': user}
                    )

                # We need to make sure Schedule objects exist for periods 1 to 8 on this weekday.
                # If they do not, we create default schedules for the class.
                schedules_by_period = {}
                for p in range(1, 9):
                    sched = Schedule.objects.filter(student_class=advised_class, day=weekday, period=p).first()
                    if not sched:
                        # Find a subject for this class, or create/use a default subject
                        from accounts.models import Subject
                        subject = Subject.objects.filter(student_class=advised_class).first()
                        if not subject:
                            subject = Subject.objects.filter(department=advised_class.department).first()
                        if not subject:
                            subject, _ = Subject.objects.get_or_create(
                                name="General",
                                code="GEN",
                                department=advised_class.department
                            )
                        # Standard hour calculation
                        start_hour = 9 + (p - 1)
                        if p >= 5:
                            start_hour += 1 # lunch break
                        start_time = datetime.time(start_hour, 0)
                        end_time = datetime.time(start_hour + 1, 0)
                        sched = Schedule.objects.create(
                            student_class=advised_class,
                            subject=subject,
                            period=p,
                            day=weekday,
                            start_time=start_time,
                            end_time=end_time
                        )
                    schedules_by_period[p] = sched

                # Update or create attendance records for each student in the class
                students = Student.objects.filter(student_class=advised_class)
                updated_records_count = 0
                
                for student in students:
                    student_payload = attendance_data.get(str(student.user_id)) or attendance_data.get(student.user_id)
                    # If student is not in payload, they default to all present
                    if not student_payload:
                        student_payload = {
                            'overall_status': 'Present',
                            'periods': {}
                        }
                    
                    overall_status = student_payload.get('overall_status', 'Present')
                    period_statuses = student_payload.get('periods', {})
                    
                    for p in range(1, 9):
                        # Determine status for this period
                        if overall_status == 'Present':
                            p_status = 'Present'
                        elif overall_status == 'Absent':
                            p_status = 'Absent'
                        elif overall_status == 'OD':
                            p_status = 'OD'
                        elif overall_status == 'Half Day (FN Present / AN Absent)':
                            p_status = 'Present' if p <= 4 else 'Absent'
                        elif overall_status == 'Half Day (FN Absent / AN Present)':
                            p_status = 'Absent' if p <= 4 else 'Present'
                        else: # Custom
                            p_status = period_statuses.get(str(p)) or period_statuses.get(p) or 'Present'
                            
                        if p_status not in ['Present', 'Absent', 'OD', 'Leave']:
                            p_status = 'Present'
                            
                        # Save/Update
                        Attendance.objects.update_or_create(
                            student=student,
                            schedule=schedules_by_period[p],
                            date=target_date,
                            defaults={'status': p_status}
                        )
                        updated_records_count += 1
                        
                return Response({
                    'success': True,
                    'detail': f'Successfully updated daily attendance for {students.count()} students ({updated_records_count} period records).'
                })
        except Exception as e:
            return Response({'detail': f'Error saving manual attendance: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'], url_path='staff-marked-history')
    def staff_marked_history(self, request):
        user = request.user
        if user.role != 'staff' and user.role != 'hod':
            return Response({'detail': 'Only staff members can view marked attendance history.'}, status=status.HTTP_403_FORBIDDEN)
        
        from .models import PeriodLock, Attendance
        from timetable.models import Schedule
        from accounts.models import Student

        # Fetch locks where staff = user
        locks = list(PeriodLock.objects.filter(staff=user).select_related('student_class').order_by('-date', 'period'))
        
        class_ids = set(lock.student_class_id for lock in locks)
        dates = set(lock.date for lock in locks)
        
        # 1. Fetch schedules for these classes in batch
        schedules = Schedule.objects.filter(student_class_id__in=class_ids).select_related('subject')
        schedule_map = {}
        for s in schedules:
            schedule_map[(s.student_class_id, s.day, s.period)] = s
            
        # 2. Fetch all student counts for these classes in batch
        from django.db.models import Count
        class_student_counts = dict(Student.objects.filter(
            student_class_id__in=class_ids
        ).values('student_class_id').annotate(count=Count('user_id')).values_list('student_class_id', 'count'))
        
        # 3. Fetch all attendance statuses in batch
        attendances = Attendance.objects.filter(
            student__student_class_id__in=class_ids,
            date__in=dates
        ).select_related('schedule__subject')
        
        # Group by (class_id, date, period) -> list of statuses
        from collections import defaultdict
        att_status_map = defaultdict(list)
        for att in attendances:
            period = att.schedule.period if att.schedule else None
            if period:
                att_status_map[(att.student.student_class_id, att.date, period)].append(att.status)
                
        # Group sessions by date string
        history_by_date = {}
        
        for lock in locks:
            d_str = lock.date.strftime('%Y-%m-%d')
            if d_str not in history_by_date:
                history_by_date[d_str] = {
                    'date': d_str,
                    'formatted_date': lock.date.strftime('%A, %d %B %Y'),
                    'sessions': []
                }
            
            weekday = lock.date.strftime('%A')
            sched = schedule_map.get((lock.student_class_id, weekday, lock.period))
            
            if not sched:
                # Fallback to check if any attendance has schedule
                statuses_records = Attendance.objects.filter(
                    student__student_class=lock.student_class,
                    date=lock.date,
                    schedule__period=lock.period
                ).select_related('schedule__subject').first()
                if statuses_records:
                    sched = statuses_records.schedule

            subject_name = sched.subject.name if (sched and sched.subject) else "General"
            subject_code = sched.subject.code if (sched and sched.subject) else "GEN"

            # Get statuses from pre-fetched map
            statuses = att_status_map.get((lock.student_class_id, lock.date, lock.period), [])
            total_students = class_student_counts.get(lock.student_class_id, 0)
            
            present_count = statuses.count('Present')
            absent_count = statuses.count('Absent')
            od_count = statuses.count('OD')
            leave_count = statuses.count('Leave')

            if total_students == 0 and statuses:
                total_students = len(statuses)

            history_by_date[d_str]['sessions'].append({
                'period': lock.period,
                'class_id': lock.student_class.id,
                'class_name': lock.student_class.name,
                'subject_code': subject_code,
                'subject_name': subject_name,
                'present_count': present_count,
                'absent_count': absent_count,
                'od_count': od_count,
                'leave_count': leave_count,
                'total_students': total_students,
                'marked_at': lock.date.strftime('%Y-%m-%d')
            })

        # Transform into list and calculate S.No (1 to max 8) for each date's sessions
        history_list = []
        for d_str, data in history_by_date.items():
            sorted_sessions = sorted(data['sessions'], key=lambda x: x['period'])
            # Ensure period limit max 8 periods
            sessions_capped = sorted_sessions[:8]
            for index, session in enumerate(sessions_capped, start=1):
                session['s_no'] = index
                
            data['sessions'] = sessions_capped
            data['total_sessions'] = len(sessions_capped)
            history_list.append(data)

        return Response({'history': history_list})

    @action(detail=False, methods=['get'], url_path='staff-history-session-detail')
    def staff_history_session_detail(self, request):
        user = request.user
        if user.role != 'staff' and user.role != 'hod':
            return Response({'detail': 'Only staff members can view session details.'}, status=status.HTTP_403_FORBIDDEN)

        date_str = request.query_params.get('date')
        class_id = request.query_params.get('class_id')
        period_val = request.query_params.get('period')

        if not date_str or not class_id or not period_val:
            return Response({'detail': 'Missing required query parameters: date, class_id, period.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            target_date = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()
            period_num = int(period_val)
        except ValueError:
            return Response({'detail': 'Invalid date format or period number.'}, status=status.HTTP_400_BAD_REQUEST)

        from accounts.models import Class, Student
        from timetable.models import Schedule

        student_class = get_object_or_404(Class, pk=class_id)
        weekday = target_date.strftime('%A')

        sched = Schedule.objects.filter(student_class=student_class, period=period_num, day=weekday).select_related('subject').first()
        
        subject_name = sched.subject.name if (sched and sched.subject) else "General"
        subject_code = sched.subject.code if (sched and sched.subject) else "GEN"

        students = Student.objects.filter(student_class=student_class).select_related('user').order_by('reg_no', 'user__username')

        atts = Attendance.objects.filter(
            student__student_class=student_class,
            date=target_date,
            schedule__period=period_num
        )
        att_dict = {a.student_id: a.status for a in atts}

        student_list = []
        for idx, s in enumerate(students, start=1):
            status_val = att_dict.get(s.user_id, 'Absent')
            student_list.append({
                's_no': idx,
                'student_id': s.user_id,
                'db_student_id': s.user_id,
                'roll_no': s.roll_no or '',
                'reg_no': s.reg_no or s.user.username,
                'name': f"{s.user.first_name} {s.user.last_name}".strip() or s.user.username,
                'status': status_val
            })

        return Response({
            'date': date_str,
            'formatted_date': target_date.strftime('%A, %d %B %Y'),
            'class_id': student_class.id,
            'class_name': student_class.name,
            'period': period_num,
            'subject_code': subject_code,
            'subject_name': subject_name,
            'students': student_list
        })

    @action(detail=False, methods=['post'], url_path='update-staff-history-session')
    def update_staff_history_session(self, request):
        user = request.user
        if user.role != 'staff' and user.role != 'hod':
            return Response({'detail': 'Only staff members can update session attendance.'}, status=status.HTTP_403_FORBIDDEN)

        date_str = request.data.get('date')
        class_id = request.data.get('class_id')
        period_val = request.data.get('period')
        statuses = request.data.get('statuses', {})

        if not date_str or not class_id or period_val is None:
            return Response({'detail': 'Missing date, class_id, or period.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            target_date = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()
            period_num = int(period_val)
            weekday = target_date.strftime('%A')
        except ValueError:
            return Response({'detail': 'Invalid date format or period number.'}, status=status.HTTP_400_BAD_REQUEST)

        from accounts.models import Class, Student
        from timetable.models import Schedule
        from .models import PeriodLock

        student_class = get_object_or_404(Class, pk=class_id)

        # Check or update lock to staff user
        PeriodLock.objects.update_or_create(
            student_class=student_class,
            date=target_date,
            period=period_num,
            defaults={'staff': user}
        )

        sched = Schedule.objects.filter(student_class=student_class, period=period_num, day=weekday).first()
        if not sched:
            from accounts.models import Subject
            subject = Subject.objects.filter(student_class=student_class).first() or Subject.objects.filter(department=student_class.department).first()
            if not subject:
                subject, _ = Subject.objects.get_or_create(name="General", code="GEN", department=student_class.department)
            
            start_hour = 9 + (period_num - 1)
            if period_num >= 5:
                start_hour += 1
            sched = Schedule.objects.create(
                student_class=student_class,
                subject=subject,
                period=period_num,
                day=weekday,
                start_time=datetime.time(start_hour, 0),
                end_time=datetime.time(start_hour + 1, 0)
            )

        students = Student.objects.filter(student_class=student_class)
        from django.db import transaction
        with transaction.atomic():
            for student in students:
                st_status = (
                    statuses.get(str(student.user_id)) or
                    statuses.get(str(student.id)) or
                    statuses.get(student.user_id) or
                    statuses.get(student.id)
                )
                if st_status and st_status in ['Present', 'Absent', 'OD', 'Leave']:
                    Attendance.objects.update_or_create(
                        student=student,
                        schedule=sched,
                        date=target_date,
                        defaults={'status': st_status}
                    )

        return Response({
            'success': True,
            'detail': f'Session attendance updated successfully for Class {student_class.name}, Period {period_num}.'
        })

