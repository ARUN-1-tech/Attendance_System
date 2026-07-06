from django.test import TestCase
from django.core.files.uploadedfile import SimpleUploadedFile
from accounts.models import User, Department, Class, Student, Staff, Subject
import csv
import io
import openpyxl

class StudentBulkImportTestCase(TestCase):
    def setUp(self):
        # Create Department
        self.dept = Department.objects.create(name='Computer Science')
        
        # Create Class
        self.clazz = Class.objects.create(name='B.Tech CS', year=3, section='A', department=self.dept)
        
        # Create tutor users
        self.tutor1_user = User.objects.create_user('tutor1', 'tutor1@example.com', 'pass123')
        self.tutor1_user.role = 'staff'
        self.tutor1_user.save()
        Staff.objects.create(user=self.tutor1_user, staff_type='Tutor')
        
        self.tutor2_user = User.objects.create_user('tutor2', 'tutor2@example.com', 'pass123')
        self.tutor2_user.role = 'staff'
        self.tutor2_user.save()
        Staff.objects.create(user=self.tutor2_user, staff_type='Tutor')
        
        self.tutor3_user = User.objects.create_user('tutor3', 'tutor3@example.com', 'pass123')
        self.tutor3_user.role = 'staff'
        self.tutor3_user.save()
        Staff.objects.create(user=self.tutor3_user, staff_type='Tutor')
        
        # Assign tutors to class
        self.clazz.tutor1 = self.tutor1_user
        self.clazz.tutor2 = self.tutor2_user
        self.clazz.tutor3 = self.tutor3_user
        self.clazz.save()
        
        # Create Advisor (importer user)
        self.advisor_user = User.objects.create_user('advisor1', 'advisor1@example.com', 'advisor123')
        self.advisor_user.role = 'staff'
        self.advisor_user.department = self.dept
        self.advisor_user.save()
        self.advisor_staff = Staff.objects.create(user=self.advisor_user, staff_type='Advisor')
        
        # Create an existing student to trigger "Username already exists" error
        self.existing_user = User.objects.create_user('student_existing', 'existing@example.com', 'pass123')
        self.existing_user.role = 'student'
        self.existing_user.department = self.dept
        self.existing_user.save()
        self.existing_student = Student.objects.create(user=self.existing_user, student_class=self.clazz, reg_no='REG_EXISTING')
        
        # Log in as Advisor
        self.client.login(username='advisor1', password='advisor123')

    def test_invalid_file_type_rejected(self):
        file_content = b"some dummy txt content"
        uploaded_file = SimpleUploadedFile("students.txt", file_content, content_type="text/plain")
        
        response = self.client.post('/api/students/bulk_create/', {'file': uploaded_file})
        self.assertEqual(response.status_code, 400)
        self.assertIn("Please upload a CSV (.csv) or Excel (.xlsx, .xls) file.", response.data['detail'])

    def test_missing_required_columns(self):
        csv_buffer = io.StringIO()
        writer = csv.writer(csv_buffer)
        writer.writerow(['username', 'password', 'class']) # missing 'year'
        writer.writerow(['test_user', 'pass123', 'B.Tech CS'])
        
        uploaded_file = SimpleUploadedFile("students.csv", csv_buffer.getvalue().encode('utf-8'), content_type="text/csv")
        response = self.client.post('/api/students/bulk_create/', {'file': uploaded_file})
        self.assertEqual(response.status_code, 400)
        self.assertIn("Missing required CSV columns", response.data['detail'])
        self.assertIn("year", response.data['detail'])

    def test_csv_bulk_import_63_students(self):
        csv_buffer = io.StringIO()
        writer = csv.writer(csv_buffer)
        writer.writerow(['username', 'password', 'class', 'year', 'register_no', 'roll_no'])
        
        for i in range(1, 64): # 1 to 63
            writer.writerow([
                f'csv_student_{i:02d}',
                'password123',
                'B.Tech CS',
                '3',
                f'REG_CSV_{i:03d}',
                f'ROLL_CSV_{i:03d}'
            ])
            
        uploaded_file = SimpleUploadedFile("students.csv", csv_buffer.getvalue().encode('utf-8'), content_type="text/csv")
        response = self.client.post('/api/students/bulk_create/', {'file': uploaded_file})
        
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.data['success'])
        self.assertEqual(response.data['created'], 63)
        self.assertEqual(response.data['failed'], 0)
        self.assertEqual(len(response.data['errors']), 0)
        
        # Verify summary counts
        self.assertEqual(response.data['summary']['imported'], 63)
        self.assertEqual(response.data['summary']['tutor1'], 21)
        self.assertEqual(response.data['summary']['tutor2'], 21)
        self.assertEqual(response.data['summary']['tutor3'], 21)
        
        # Verify tutor assignment in DB
        self.assertEqual(Student.objects.filter(user__username__startswith='csv_student', tutor=self.tutor1_user).count(), 21)
        self.assertEqual(Student.objects.filter(user__username__startswith='csv_student', tutor=self.tutor2_user).count(), 21)
        self.assertEqual(Student.objects.filter(user__username__startswith='csv_student', tutor=self.tutor3_user).count(), 21)
        
        # Verify advisor fallback (tutor3 since clazz.advisor is None)
        self.assertEqual(Student.objects.filter(user__username__startswith='csv_student', advisor=self.tutor3_user).count(), 63)

    def test_excel_bulk_import_64_students(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['username', 'password', 'class', 'year', 'register_no', 'roll_no'])
        
        for i in range(1, 65): # 1 to 64
            ws.append([
                f'xlsx_student_{i:02d}',
                'password123',
                'B.Tech CS',
                3,
                f'REG_XLSX_{i:03d}',
                f'ROLL_XLSX_{i:03d}'
            ])
            
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        uploaded_file = SimpleUploadedFile("students.xlsx", excel_buffer.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response = self.client.post('/api/students/bulk_create/', {'file': uploaded_file})
        
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.data['success'])
        self.assertEqual(response.data['created'], 64)
        
        # Verify 64 -> 22, 21, 21
        self.assertEqual(response.data['summary']['tutor1'], 22)
        self.assertEqual(response.data['summary']['tutor2'], 21)
        self.assertEqual(response.data['summary']['tutor3'], 21)

    def test_excel_bulk_import_65_students(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['username', 'password', 'class', 'year', 'register_no', 'roll_no'])
        
        for i in range(1, 66): # 1 to 65
            ws.append([
                f'xlsx_student2_{i:02d}',
                'password123',
                'B.Tech CS',
                3,
                f'REG_XLSX2_{i:03d}',
                f'ROLL_XLSX2_{i:03d}'
            ])
            
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        uploaded_file = SimpleUploadedFile("students.xlsx", excel_buffer.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response = self.client.post('/api/students/bulk_create/', {'file': uploaded_file})
        
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.data['success'])
        self.assertEqual(response.data['created'], 65)
        
        # Verify 65 -> 22, 22, 21
        self.assertEqual(response.data['summary']['tutor1'], 22)
        self.assertEqual(response.data['summary']['tutor2'], 22)
        self.assertEqual(response.data['summary']['tutor3'], 21)

    def test_partial_failures_and_row_wise_errors(self):
        csv_buffer = io.StringIO()
        writer = csv.writer(csv_buffer)
        writer.writerow(['username', 'password', 'class', 'year', 'register_no', 'roll_no'])
        
        # Row 2: Valid student
        writer.writerow(['new_student_val', 'password123', 'B.Tech CS', '3', 'REG_NEW', 'ROLL_NEW'])
        # Row 3: Existing username
        writer.writerow(['student_existing', 'password123', 'B.Tech CS', '3', 'REG_OTHER', 'ROLL_OTHER'])
        # Row 4: Completely empty row (should be skipped)
        writer.writerow(['', '', '', '', '', ''])
        # Row 5: Invalid class (year 4 doesn't exist)
        writer.writerow(['new_student_val2', 'password123', 'B.Tech CS', '4', 'REG_NEW2', 'ROLL_NEW2'])
        
        uploaded_file = SimpleUploadedFile("students.csv", csv_buffer.getvalue().encode('utf-8'), content_type="text/csv")
        response = self.client.post('/api/students/bulk_create/', {'file': uploaded_file})
        
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.data['success'])
        self.assertEqual(response.data['created'], 1)
        self.assertEqual(response.data['failed'], 2)
        
        errors = response.data['errors']
        self.assertEqual(len(errors), 2)
        self.assertIn("Row 3 : Username already exists", errors)
        self.assertIn("Row 5 : Class not found", errors)

    def test_tolerant_class_lookup(self):
        # Create classes with different formats to test tolerant class lookup
        Class.objects.create(name='B.E CSE', year=3, section='A', department=self.dept, tutor3=self.advisor_user)
        
        # Test class matching for: B.E. CSE A, B.E CSE_A, BE CSE A, b.e cse a
        csv_buffer = io.StringIO()
        writer = csv.writer(csv_buffer)
        writer.writerow(['username', 'password', 'class', 'year', 'register_no'])
        writer.writerow(['student_class_test1', 'pass123', 'B.E. CSE A', '3', 'REG_CLASS1'])
        writer.writerow(['student_class_test2', 'pass123', 'B.E CSE_A', '3', 'REG_CLASS2'])
        writer.writerow(['student_class_test3', 'pass123', 'BE CSE A', '3', 'REG_CLASS3'])
        
        uploaded_file = SimpleUploadedFile("students.csv", csv_buffer.getvalue().encode('utf-8'), content_type="text/csv")
        response = self.client.post('/api/students/bulk_create/', {'file': uploaded_file})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['created'], 3)
        self.assertEqual(response.data['failed'], 0)

    def test_department_missing_error(self):
        # Log in as a user with no department assigned
        advisor_no_dept = User.objects.create_user('advisor_no_dept', 'no_dept@example.com', 'pass123')
        advisor_no_dept.role = 'staff'
        advisor_no_dept.save()
        Staff.objects.create(user=advisor_no_dept, staff_type='Advisor')
        
        self.client.login(username='advisor_no_dept', password='pass123')
        
        csv_buffer = io.StringIO()
        writer = csv.writer(csv_buffer)
        writer.writerow(['username', 'password', 'class', 'year'])
        writer.writerow(['student_no_dept', 'pass123', 'B.Tech CS', '3'])
        
        uploaded_file = SimpleUploadedFile("students.csv", csv_buffer.getvalue().encode('utf-8'), content_type="text/csv")
        response = self.client.post('/api/students/bulk_create/', {'file': uploaded_file})
        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.data['detail'], "HOD/Advisor has no department assigned.")

    def test_duplicate_checks_reg_no_and_roll_no(self):
        # Register an existing student with a roll number to test DB duplicates
        self.existing_student.roll_no = 'ROLL_EXISTING'
        self.existing_student.save()
        
        csv_buffer = io.StringIO()
        writer = csv.writer(csv_buffer)
        writer.writerow(['username', 'password', 'class', 'year', 'register_no', 'roll_no'])
        
        # Row 2: Duplicate registration number in DB
        writer.writerow(['student_dup1', 'pass123', 'B.Tech CS', '3', 'REG_EXISTING', 'ROLL_NEW1'])
        # Row 3: Duplicate roll number in DB
        writer.writerow(['student_dup2', 'pass123', 'B.Tech CS', '3', 'REG_NEW1', 'ROLL_EXISTING'])
        # Row 4: Duplicate registration number in the same file (first one succeeded)
        writer.writerow(['student_dup3', 'pass123', 'B.Tech CS', '3', 'REG_FILE_DUP', 'ROLL_NEW2'])
        # Row 5: Duplicate registration number in the same file (this one fails)
        writer.writerow(['student_dup4', 'pass123', 'B.Tech CS', '3', 'REG_FILE_DUP', 'ROLL_NEW3'])
        # Row 6: Duplicate roll number in the same file (first one succeeded)
        writer.writerow(['student_dup5', 'pass123', 'B.Tech CS', '3', 'REG_NEW4', 'ROLL_FILE_DUP'])
        # Row 7: Duplicate roll number in the same file (this one fails)
        writer.writerow(['student_dup6', 'pass123', 'B.Tech CS', '3', 'REG_NEW5', 'ROLL_FILE_DUP'])
        
        uploaded_file = SimpleUploadedFile("students.csv", csv_buffer.getvalue().encode('utf-8'), content_type="text/csv")
        response = self.client.post('/api/students/bulk_create/', {'file': uploaded_file})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['created'], 2)
        self.assertEqual(response.data['failed'], 4)
        
        errors = response.data['errors']
        self.assertIn("Row 2 : Registration number already exists", errors)
        self.assertIn("Row 3 : Roll number already exists", errors)
        self.assertIn("Row 5 : Registration number already exists", errors)
        self.assertIn("Row 7 : Roll number already exists", errors)

    def test_excel_parsing_float_to_int_and_scientific(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['username', 'password', 'class', 'year', 'register_no', 'roll_no'])
        
        # Add year as float, register_no and roll_no with float/scientific representations
        ws.append(['float_student', 'pass123', 'B.Tech CS', 3.0, 241040001.0, 2.4104e+08])
        
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        uploaded_file = SimpleUploadedFile("students.xlsx", excel_buffer.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response = self.client.post('/api/students/bulk_create/', {'file': uploaded_file})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['created'], 1)
        
        student = Student.objects.get(user__username='float_student')
        self.assertEqual(student.reg_no, '241040001')
        self.assertEqual(student.roll_no, '241040000')

    def test_tutor_edit_delete_permissions(self):
        import json
        
        # Assign department to tutors so authorization passes
        self.tutor1_user.department = self.dept
        self.tutor1_user.save()
        self.tutor2_user.department = self.dept
        self.tutor2_user.save()
        self.tutor3_user.department = self.dept
        self.tutor3_user.save()
        
        # Create tutor student
        tutor_student_user = User.objects.create_user('tutored_stud', 'tutored@example.com', 'pass123')
        tutor_student_user.role = 'student'
        tutor_student_user.department = self.dept
        tutor_student_user.save()
        tutor_student = Student.objects.create(user=tutor_student_user, student_class=self.clazz, reg_no='REG_TUTORED')
        
        # Create other student
        other_student_user = User.objects.create_user('other_stud', 'other@example.com', 'pass123')
        other_student_user.role = 'student'
        other_student_user.department = self.dept
        other_student_user.save()
        other_student = Student.objects.create(user=other_student_user, student_class=self.clazz, reg_no='REG_OTHER')
        
        # Refresh from DB to see actual tutors assigned by auto_assign_tutors
        tutor_student.refresh_from_db()
        other_student.refresh_from_db()
        
        assigned_tutor_user = tutor_student.tutor
        other_tutor_user = self.tutor1_user if assigned_tutor_user != self.tutor1_user else self.tutor2_user
        
        # Log in as the assigned tutor
        self.client.login(username=assigned_tutor_user.username, password='pass123')
        
        # Attempt to edit tutored student (should succeed)
        response = self.client.put(f'/api/students/{tutor_student_user.id}/', data=json.dumps({
            'user': {
                'username': 'tutored_stud',
                'email': 'tutored_new@example.com',
                'first_name': 'tutored_stud',
                'role': 'student',
                'password': 'pass123'
            },
            'student_class': self.clazz.id,
            'reg_no': 'REG_TUTORED_NEW'
        }), content_type='application/json')
        self.assertEqual(response.status_code, 200)
        
        # Log in as the other tutor
        self.client.login(username=other_tutor_user.username, password='pass123')
        
        # Attempt to edit tutored student as other tutor (should fail with 403)
        response = self.client.put(f'/api/students/{tutor_student_user.id}/', data=json.dumps({
            'user': {
                'username': 'tutored_stud',
                'email': 'tutored_new@example.com',
                'first_name': 'tutored_stud',
                'role': 'student',
                'password': 'pass123'
            },
            'student_class': self.clazz.id,
            'reg_no': 'REG_TUTORED_NEW'
        }), content_type='application/json')
        self.assertEqual(response.status_code, 403)
        
        # Attempt to delete tutored student as other tutor (should fail with 403)
        response = self.client.delete(f'/api/students/{tutor_student_user.id}/')
        self.assertEqual(response.status_code, 403)
        
        # Log in as the assigned tutor again
        self.client.login(username=assigned_tutor_user.username, password='pass123')
        
        # Attempt to delete tutored student (should succeed)
        response = self.client.delete(f'/api/students/{tutor_student_user.id}/')
        self.assertEqual(response.status_code, 204)

class UserPasswordChangeAndManualAttendanceTestCase(TestCase):
    def setUp(self):
        self.dept = Department.objects.create(name='Computer Science')
        self.clazz = Class.objects.create(name='B.Tech CS', year=3, section='A', department=self.dept)
        
        self.staff_user = User.objects.create_user('staff_user', 'staff@example.com', 'staffpass123')
        self.staff_user.role = 'staff'
        self.staff_user.department = self.dept
        self.staff_user.save()
        self.staff = Staff.objects.create(user=self.staff_user, staff_type='Tutor')

        self.student_user = User.objects.create_user('stud_user', 'student@example.com', 'studpass123')
        self.student_user.role = 'student'
        self.student_user.department = self.dept
        self.student_user.save()
        self.student = Student.objects.create(user=self.student_user, student_class=self.clazz, reg_no='REG_123', tutor=self.staff_user)

        self.subject = Subject.objects.create(name='Automata Theory', code='CS301', department=self.dept, student_class=self.clazz)
        
    def test_change_password_success(self):
        self.client.login(username='staff_user', password='staffpass123')
        response = self.client.post(f'/api/users/{self.staff_user.id}/change_password/', {
            'current_password': 'staffpass123',
            'new_password': 'newstaffpass123'
        })
        self.assertEqual(response.status_code, 200)
        
        # Verify password actually updated
        self.client.logout()
        login_success = self.client.login(username='staff_user', password='newstaffpass123')
        self.assertTrue(login_success)

    def test_change_password_invalid_current(self):
        self.client.login(username='staff_user', password='staffpass123')
        response = self.client.post(f'/api/users/{self.staff_user.id}/change_password/', {
            'current_password': 'wrongpassword',
            'new_password': 'newstaffpass123'
        })
        self.assertEqual(response.status_code, 400)
        self.assertIn('Incorrect current password', response.data['detail'])

    def test_change_password_unauthorized(self):
        self.client.login(username='stud_user', password='studpass123')
        # Attempt to change staff user's password
        response = self.client.post(f'/api/users/{self.staff_user.id}/change_password/', {
            'current_password': 'staffpass123',
            'new_password': 'newstaffpass123'
        })
        self.assertEqual(response.status_code, 403)

    def test_manual_class_students_fetch(self):
        self.client.login(username='staff_user', password='staffpass123')
        response = self.client.get('/api/attendances/manual-class-students/', {
            'class_id': self.clazz.id,
            'subject_id': self.subject.id,
            'date': '2026-06-25' # Thursday
        })
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data['students']), 1)
        self.assertEqual(response.data['students'][0]['id'], self.student_user.id)
        self.assertEqual(response.data['students'][0]['current_status'], 'Present')

    def test_save_class_manual_attendance(self):
        self.client.login(username='staff_user', password='staffpass123')
        response = self.client.post('/api/attendances/save-class-manual-attendance/', {
            'class_id': self.clazz.id,
            'subject_id': self.subject.id,
            'date': '2026-06-25',
            'statuses': {
                str(self.student_user.id): 'Present'
            }
        }, content_type='application/json')
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.data['success'])
        
        # Verify attendance record created
        from attendance.models import Attendance
        attendance_records = Attendance.objects.filter(student=self.student, date='2026-06-25')
        self.assertEqual(attendance_records.count(), 1)
        self.assertEqual(attendance_records.first().status, 'Present')

    def test_advisor_class_students_fetch(self):
        # Set staff_type to Advisor and make them advisor of the class
        self.staff.staff_type = 'Advisor'
        self.staff.save()
        self.clazz.advisor = self.staff_user
        self.clazz.save()
        
        self.client.login(username='staff_user', password='staffpass123')
        response = self.client.get('/api/attendances/advisor-class-students/', {
            'date': '2026-06-25' # Thursday
        })
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['class_name'], 'B.Tech CS - 3 - A')
        self.assertEqual(len(response.data['students']), 1)
        self.assertEqual(response.data['students'][0]['id'], self.student_user.id)
        # Default all periods to Present in fetched statuses if not created in DB
        self.assertEqual(response.data['students'][0]['statuses']['1'], 'Present')
        self.assertEqual(len(response.data['periods']), 8)

    def test_save_advisor_manual_attendance(self):
        self.staff.staff_type = 'Advisor'
        self.staff.save()
        self.clazz.advisor = self.staff_user
        self.clazz.save()
        
        self.client.login(username='staff_user', password='staffpass123')
        
        # Test 1: Present (Full Day)
        response = self.client.post('/api/attendances/save-advisor-manual-attendance/', {
            'date': '2026-06-25',
            'attendance_data': {
                str(self.student_user.id): {
                    'overall_status': 'Present',
                    'periods': {}
                }
            }
        }, content_type='application/json')
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.data['success'])
        
        from attendance.models import Attendance
        self.assertEqual(Attendance.objects.filter(student=self.student, date='2026-06-25').count(), 8)
        self.assertTrue(all(att.status == 'Present' for att in Attendance.objects.filter(student=self.student, date='2026-06-25')))

        # Test 2: Half Day (FN Present / AN Absent)
        response = self.client.post('/api/attendances/save-advisor-manual-attendance/', {
            'date': '2026-06-26',
            'attendance_data': {
                str(self.student_user.id): {
                    'overall_status': 'Half Day (FN Present / AN Absent)',
                    'periods': {}
                }
            }
        }, content_type='application/json')
        self.assertEqual(response.status_code, 200)
        
        atts_fn = Attendance.objects.filter(student=self.student, date='2026-06-26', schedule__period__lte=4)
        atts_an = Attendance.objects.filter(student=self.student, date='2026-06-26', schedule__period__gt=4)
        self.assertEqual(atts_fn.count(), 4)
        self.assertEqual(atts_an.count(), 4)
        self.assertTrue(all(att.status == 'Present' for att in atts_fn))
        self.assertTrue(all(att.status == 'Absent' for att in atts_an))

        # Test 3: Custom
        response = self.client.post('/api/attendances/save-advisor-manual-attendance/', {
            'date': '2026-06-27',
            'attendance_data': {
                str(self.student_user.id): {
                    'overall_status': 'Custom',
                    'periods': {
                        '1': 'Present',
                        '2': 'Absent',
                        '3': 'OD',
                        '4': 'Leave',
                        '5': 'Present',
                        '6': 'Absent',
                        '7': 'Present'
                    }
                }
            }
        }, content_type='application/json')
        self.assertEqual(response.status_code, 200)
        
        self.assertEqual(Attendance.objects.get(student=self.student, date='2026-06-27', schedule__period=1).status, 'Present')
        self.assertEqual(Attendance.objects.get(student=self.student, date='2026-06-27', schedule__period=2).status, 'Absent')
        self.assertEqual(Attendance.objects.get(student=self.student, date='2026-06-27', schedule__period=3).status, 'OD')
        self.assertEqual(Attendance.objects.get(student=self.student, date='2026-06-27', schedule__period=4).status, 'Leave')

    def test_daily_report_half_day_aggregation(self):
        self.staff.staff_type = 'Advisor'
        self.staff.save()
        self.clazz.advisor = self.staff_user
        self.clazz.save()
        
        # Save a mix of Present/Absent for June 26
        self.client.login(username='staff_user', password='staffpass123')
        self.client.post('/api/attendances/save-advisor-manual-attendance/', {
            'date': '2026-06-26',
            'attendance_data': {
                str(self.student_user.id): {
                    'overall_status': 'Half Day (FN Present / AN Absent)',
                    'periods': {}
                }
            }
        }, content_type='application/json')

        # Query reports API
        response = self.client.get('/api/attendance/reports/', {
            'report_type': 'class',
            'class_id': self.clazz.id,
            'report_mode': 'day',
            'date': '2026-06-26'
        })
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data), 1)
        self.assertEqual(response.data[0]['status'], 'Half Day')

    def test_student_stats_returns_all_subjects_even_with_no_attendance(self):
        from accounts.models import Subject
        # Create a subject assigned to class but with no schedules or attendance
        new_sub = Subject.objects.create(
            name='Physics',
            code='PHY101',
            student_class=self.clazz,
            department=self.dept
        )
        
        self.client.login(username='stud_user', password='studpass123')
        response = self.client.get('/api/attendance/student-stats/stud_user/')
        self.assertEqual(response.status_code, 200)
        
        subjects = response.data['subjects_breakdown']
        physics_sub = next((s for s in subjects if s['code'] == 'PHY101'), None)
        self.assertIsNotNone(physics_sub)
        self.assertEqual(physics_sub['percentage'], 100.0)
        self.assertEqual(physics_sub['total_periods'], 0)

    def test_bulk_import_subjects_success(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        from accounts.models import Subject
        
        self.staff.staff_type = 'Advisor'
        self.staff.save()
        self.clazz.advisor = self.staff_user
        self.clazz.save()
        
        self.client.login(username='staff_user', password='staffpass123')
        
        csv_content = b"code,name\nPHY101,Physics\nCHM101,Chemistry"
        uploaded_file = SimpleUploadedFile("subjects.csv", csv_content, content_type="text/csv")
        
        response = self.client.post('/api/subjects/bulk-import/', {
            'file': uploaded_file
        })
        
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.data['success'])
        self.assertEqual(response.data['created'], 2)
        
        # Verify database
        self.assertTrue(Subject.objects.filter(code='PHY101', student_class=self.clazz).exists())
        self.assertTrue(Subject.objects.filter(code='CHM101', student_class=self.clazz).exists())

    def test_period_lock_conflict(self):
        from attendance.models import PeriodLock
        # Create user B
        staff2_user = User.objects.create_user('staff2', 'staff2@example.com', 'staffpass123')
        staff2_user.role = 'staff'
        staff2_user.save()
        Staff.objects.create(user=staff2_user, staff_type='Normal')

        # Create period lock by staff2 for clazz today (Thursday is weekday for target_date)
        from django.utils import timezone
        import datetime
        target_date = datetime.date(2026, 6, 25) # Thursday
        PeriodLock.objects.create(
            student_class=self.clazz,
            date=target_date,
            period=1,
            staff=staff2_user
        )

        # Login as staff_user (owner of clazz advisor is staff_user if set, but we login as staff_user here)
        self.client.login(username='staff_user', password='staffpass123')

        # Attempt to save subject manual attendance for locked period 1 -> should fail
        response = self.client.post('/api/attendances/save-class-manual-attendance/', {
            'class_id': self.clazz.id,
            'subject_id': self.subject.id,
            'date': '2026-06-25',
            'period': '1',
            'statuses': {
                str(self.student_user.id): 'Present'
            }
        }, content_type='application/json')
        self.assertEqual(response.status_code, 400)
        self.assertIn("already marked/used", response.data['detail'])

        # Attempt to save advisor manual attendance for locked period 1 -> should succeed (advisor overrides locks)
        self.staff.staff_type = 'Advisor'
        self.staff.save()
        self.clazz.advisor = self.staff_user
        self.clazz.save()
        response = self.client.post('/api/attendances/save-advisor-manual-attendance/', {
            'date': '2026-06-25',
            'attendance_data': {
                str(self.student_user.id): {
                    'overall_status': 'Present',
                    'periods': {}
                }
            }
        }, content_type='application/json')
        self.assertEqual(response.status_code, 200)

        # But if advisor tries to save subject manual attendance for ANY locked period -> should fail (advisors must use whole day page)
        # Verify that lock is now owned by advisor (staff_user)
        self.assertEqual(PeriodLock.objects.get(student_class=self.clazz, date='2026-06-25', period=1).staff, self.staff_user)
        
        response = self.client.post('/api/attendances/save-class-manual-attendance/', {
            'class_id': self.clazz.id,
            'subject_id': self.subject.id,
            'date': '2026-06-25',
            'period': '1',
            'statuses': {
                str(self.student_user.id): 'Present'
            }
        }, content_type='application/json')
        self.assertEqual(response.status_code, 400)
        self.assertIn("must edit it through the Advisor Whole Day", response.data['detail'])

    def test_default_present_in_subject_manual_save(self):
        self.client.login(username='staff_user', password='staffpass123')
        # Create an attendance record and verify default Present status
        response = self.client.post('/api/attendances/save-class-manual-attendance/', {
            'class_id': self.clazz.id,
            'subject_id': self.subject.id,
            'date': '2026-06-25',
            'period': '2', # lock period 2
            'statuses': {} # empty statuses, should default student_user to Present
        }, content_type='application/json')
        self.assertEqual(response.status_code, 200)
        
        from attendance.models import Attendance
        attendance = Attendance.objects.filter(student=self.student, date='2026-06-25', schedule__period=2).first()
        self.assertIsNotNone(attendance)
        self.assertEqual(attendance.status, 'Present')

    def test_optional_eighth_period_calculation(self):
        # We need a schedule for period 8
        from timetable.models import Schedule
        import datetime
        from attendance.models import Attendance
        
        # Create a period 1 schedule
        sched1 = Schedule.objects.create(
            student_class=self.clazz,
            subject=self.subject,
            period=1,
            day='Thursday',
            start_time=datetime.time(9, 0),
            end_time=datetime.time(10, 0)
        )
        
        # Create a period 8 schedule
        sched8 = Schedule.objects.create(
            student_class=self.clazz,
            subject=self.subject,
            period=8,
            day='Thursday',
            start_time=datetime.time(16, 0),
            end_time=datetime.time(17, 0)
        )
        
        # Login and check student stats initially - should be 0 total periods since no attendance exists
        self.client.login(username='stud_user', password='studpass123')
        response = self.client.get('/api/attendance/student-stats/stud_user/')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['total_periods'], 0)
        
        # Mark period 1 as Absent
        Attendance.objects.create(student=self.student, schedule=sched1, date='2026-06-25', status='Absent')
        response = self.client.get('/api/attendance/student-stats/stud_user/')
        self.assertEqual(response.data['total_periods'], 1)
        self.assertEqual(response.data['percentage'], 0.0)
        
        # Mark period 8 as Absent - total_periods should STILL be 1 because optional 8th period is ignored when not Present
        att8 = Attendance.objects.create(student=self.student, schedule=sched8, date='2026-06-25', status='Absent')
        response = self.client.get('/api/attendance/student-stats/stud_user/')
        self.assertEqual(response.data['total_periods'], 1)
        self.assertEqual(response.data['percentage'], 0.0)
        
        # Change period 8 to Present - total_periods should become 2, present_periods = 1, percentage = 50.0
        att8.status = 'Present'
        att8.save()
        response = self.client.get('/api/attendance/student-stats/stud_user/')
        self.assertEqual(response.data['total_periods'], 2)
        self.assertEqual(response.data['percentage'], 50.0)

    def test_subject_detail_and_csv(self):
        # We need a schedule and attendance record
        from attendance.models import Attendance
        from timetable.models import Schedule
        import datetime
        
        sched = Schedule.objects.create(
            student_class=self.clazz,
            subject=self.subject,
            period=2,
            day='Thursday',
            start_time=datetime.time(10, 0),
            end_time=datetime.time(11, 0)
        )
        
        # Create an attendance record
        Attendance.objects.create(student=self.student, schedule=sched, date='2026-06-25', status='Present')
        
        self.client.login(username='stud_user', password='studpass123')
        
        # Test JSON endpoint
        response = self.client.get(f'/api/attendances/subject-detail/?student_username=stud_user&subject_id={self.subject.id}')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['stats']['total_hours'], 1)
        self.assertEqual(response.data['stats']['effective_present'], 1)
        self.assertEqual(response.data['records'][0]['status'], 'Present')
        
        # Test CSV download endpoint
        response = self.client.get(f'/api/attendances/subject-detail/?student_username=stud_user&subject_id={self.subject.id}&download=true')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'text/csv')
        self.assertIn("STUDENT DETAILS", response.content.decode('utf-8'))
        self.assertIn("SUBJECT DETAILS", response.content.decode('utf-8'))
        self.assertIn("ATTENDANCE SUMMARY", response.content.decode('utf-8'))

    def test_advisor_subject_report_csv(self):
        # Test downloading class-wide subject report by advisor
        self.staff.staff_type = 'Advisor'
        self.staff.save()
        self.clazz.advisor = self.staff_user
        self.clazz.save()
        
        from attendance.models import Attendance
        from timetable.models import Schedule
        import datetime
        
        sched = Schedule.objects.create(
            student_class=self.clazz,
            subject=self.subject,
            period=2,
            day='Thursday',
            start_time=datetime.time(10, 0),
            end_time=datetime.time(11, 0)
        )
        Attendance.objects.create(student=self.student, schedule=sched, date='2026-06-25', status='Present')
        
        # Log in as Advisor (staff_user)
        self.client.login(username='staff_user', password='staffpass123')
        
        # Call report download API
        response = self.client.get(f'/api/attendances/advisor-subject-report/?subject_id={self.subject.id}')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'text/csv')
        self.assertIn("Register Number", response.content.decode('utf-8'))
        self.assertIn("Name", response.content.decode('utf-8'))
        self.assertIn("Percentage", response.content.decode('utf-8'))
